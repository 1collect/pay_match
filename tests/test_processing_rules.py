import unittest
from pathlib import Path
from shutil import copyfile
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook

from processor.services.excel_processor import (
    _contains_excluded_keyword,
    _extract_dbz_keys,
    _normalize_identifier,
    _parse_date,
)
from processor.services.name_matching import (
    match_person_name,
    match_person_name_nominative,
    nominative_person_name_tokens,
)
from processor.services.ner_service import NameExtraction
from processor.services.new_format_processor import (
    NewBankStatementProcessor,
    ProcessingCancelled,
    _detail_date_for_path,
    _detail_type_for_path,
)


class FakeNameExtractor:
    def __init__(self, names_by_text: dict[str, tuple[str, ...]]) -> None:
        self.names_by_text = names_by_text

    def extract_many(self, texts, should_cancel=None):
        return [
            NameExtraction(
                self.names_by_text.get(text, ()),
                "" if self.names_by_text.get(text) else "GLiNER 2 не обнаружил допустимое ФИО",
            )
            for text in texts
        ]


class ProcessingRuleTests(unittest.TestCase):
    def test_dbz_number_accepts_one_or_two_accidental_spaces(self) -> None:
        cases = {
            "Оплата по ДБЗ: ABC12 345": "ABC12345",
            "Оплата по ДБЗ: ABC12  345": "ABC12345",
            "Оплата, договор займа № ABC 12 345": "ABC12345",
        }
        for purpose, expected in cases.items():
            with self.subTest(purpose=purpose):
                self.assertIn(expected, _extract_dbz_keys(purpose))

    def test_iso_datetime_with_space_keeps_month_and_day(self) -> None:
        self.assertEqual(str(_parse_date("2026-07-10 00:00:00.0")), "2026-07-10")

    def test_detail_type_supports_filename_prefixes(self) -> None:
        cases = {
            "CAFIRSTCOLLECTIONBURO_report.xlsx": "halyk",
            "astanaplat_weekend.xls": "astana",
            "20260715_anything.xlsx": "eurasian",
            "15631_report.xls": "kazpost",
            "5400-anything.xlsx": "kazpost",
            "15522.xlsx": "kazpost",
        }
        for filename, expected in cases.items():
            with self.subTest(filename=filename):
                self.assertEqual(_detail_type_for_path(Path(filename)), expected)

        self.assertEqual(_detail_type_for_path(Path("20261340_invalid.xlsx")), "")

    def test_detail_type_keeps_legacy_filename_keywords(self) -> None:
        cases = {
            "details_astanaplat.xlsx": "astana",
            "details_eurasian.xls": "eurasian",
            "details_kazpost.xlsx": "kazpost",
            "details_halyk.xls": "halyk",
        }
        for filename, expected in cases.items():
            with self.subTest(filename=filename):
                self.assertEqual(_detail_type_for_path(Path(filename)), expected)

    def test_detail_date_is_read_from_supported_filenames(self) -> None:
        self.assertEqual(
            str(_detail_date_for_path(Path("CAFIRSTCOLLECTIONBURO_120726.XLS"))),
            "2026-07-12",
        )
        self.assertEqual(
            str(_detail_date_for_path(Path("20260715_report.xlsx"))),
            "2026-07-15",
        )

    def test_spreadsheetml_xls_detail_is_read(self) -> None:
        with TemporaryDirectory() as directory:
            detail = Path(directory) / "CAFIRSTCOLLECTIONBURO_120726.XLS"
            detail.write_text(
                """<?xml version="1.0"?>
<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet"
 xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet">
 <Worksheet ss:Name="Реестр"><Table>
  <Row><Cell><Data ss:Type="String">РЕЕСТР ПЛАТЕЖЕЙ</Data></Cell></Row>
  <Row><Cell><Data ss:Type="String">Дата платежа</Data></Cell><Cell><Data ss:Type="String">Сумма</Data></Cell><Cell><Data ss:Type="String">ФИО заемщика</Data></Cell><Cell><Data ss:Type="String">ИИН</Data></Cell></Row>
  <Row><Cell><Data ss:Type="String">12/07/2026</Data></Cell><Cell><Data ss:Type="Number">31000</Data></Cell><Cell><Data ss:Type="String">Иванов Иван</Data></Cell><Cell><Data ss:Type="String">900101300001</Data></Cell></Row>
 </Table></Worksheet>
</Workbook>""",
                encoding="utf-8",
            )

            rows, _, _, detail_date = NewBankStatementProcessor.__new__(
                NewBankStatementProcessor
            )._read_detail_rows(detail, "halyk")

            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0].amount, 31000)
            self.assertEqual(rows[0].iin, "900101300001")
            self.assertEqual(str(detail_date), "2026-07-12")

    def test_identical_detail_file_is_skipped(self) -> None:
        with TemporaryDirectory() as directory:
            root = Path(directory)
            original = root / "5400_140726.xlsx"
            duplicate = root / "5400_140726 (1).xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["Дата платежа", "ИИН", "Сумма"])
            worksheet.append(["14.07.2026", "900101300001", 200])
            workbook.save(original)
            copyfile(original, duplicate)

            batches, warnings = NewBankStatementProcessor.__new__(
                NewBankStatementProcessor
            )._read_detail_batches([original, duplicate])

            self.assertEqual(len(batches), 1)
            self.assertEqual(next(iter(batches.values())).total_amount, 200)
            self.assertTrue(any("содержимое совпадает" in warning for warning in warnings))

    def test_daily_details_are_routed_to_matching_statement_dates(self) -> None:
        with TemporaryDirectory() as directory:
            root = Path(directory)
            statement = root / "statement.xlsx"
            reference = root / "reference.xlsx"
            first_detail = root / "20260717_first.xlsx"
            second_detail = root / "20260718_second.xlsx"
            output = root / "output"

            statement_book = Workbook()
            statement_sheet = statement_book.active
            statement_sheet.append(["Дата", "Отправитель", "Кредит", "Назначение платежа"])
            statement_sheet.append(["17.07.2026", "ЕВРАЗИЙСКИЙ БАНК", 100, "Реестр"])
            statement_sheet.append(["18.07.2026", "ЕВРАЗИЙСКИЙ БАНК", 200, "Реестр"])
            statement_book.save(statement)

            reference_book = Workbook()
            reference_sheet = reference_book.active
            reference_sheet.append(["Взыскатель", "ДБЗ", "ИИН", "ФИО"])
            reference_sheet.append(["Компания", "DBZ-1", "900101300001", "Иванов Иван"])
            reference_sheet.append(["Компания", "DBZ-2", "900101300002", "Петров Петр"])
            reference_book.save(reference)

            first_book = Workbook()
            first_sheet = first_book.active
            first_sheet.append(["ИИН", "Сумма"])
            first_sheet.append(["900101300001", 100])
            first_book.save(first_detail)

            second_book = Workbook()
            second_sheet = second_book.active
            second_sheet.append(["ИИН", "Сумма"])
            second_sheet.append(["900101300002", 200])
            second_book.save(second_detail)

            result = NewBankStatementProcessor(FakeNameExtractor({})).process_many(
                statement_paths=[statement],
                reference_path=reference,
                detail_paths=[first_detail, second_detail],
                output_dir=output,
            )

            workbook = load_workbook(result.registry_path, read_only=True, data_only=True)
            worksheet = workbook["Оплата"]
            headers = [cell.value for cell in worksheet[1]]
            rows = [
                dict(zip(headers, row))
                for row in worksheet.iter_rows(min_row=2, values_only=True)
            ]
            totals_by_date = {
                str(row["Дата"])[:10]: row["Общая задолженность"]
                for row in rows
            }
            self.assertEqual(
                totals_by_date,
                {"2026-07-17": 100, "2026-07-18": 200},
            )
            workbook.close()

    def test_total_detail_amount_replaces_borrower_name_in_result(self) -> None:
        with TemporaryDirectory() as directory:
            root = Path(directory)
            statement = root / "statement.xlsx"
            reference = root / "reference.xlsx"
            detail = root / "CAFIRSTCOLLECTIONBURO_report.xlsx"
            output = root / "output"

            statement_book = Workbook()
            statement_sheet = statement_book.active
            statement_sheet.append(["Дата", "Отправитель", "Кредит", "Назначение платежа"])
            statement_sheet.append(["15.07.2026", "Народный банк", 300, "Реестр 9005981"])
            statement_book.save(statement)

            reference_book = Workbook()
            reference_sheet = reference_book.active
            reference_sheet.append(["Взыскатель", "ДБЗ", "ИИН", "ФИО"])
            reference_sheet.append(["Компания", "DBZ-1", "900101300001", "Иванов Иван"])
            reference_sheet.append(["Компания", "DBZ-2", "900101300002", "Петров Петр"])
            reference_book.save(reference)

            detail_book = Workbook()
            detail_sheet = detail_book.active
            detail_sheet.append(["ИИН", "ФИО заемщика", "Сумма"])
            detail_sheet.append(["900101300001", "Иванов Иван", 100])
            detail_sheet.append(["900101300002", "Петров Петр", 200])
            detail_book.save(detail)

            result = NewBankStatementProcessor(FakeNameExtractor({})).process_many(
                statement_paths=[statement],
                reference_path=reference,
                detail_paths=[detail],
                output_dir=output,
            )

            workbook = load_workbook(result.registry_path, read_only=True, data_only=True)
            worksheet = workbook["Оплата"]
            headers = [cell.value for cell in worksheet[1]]
            rows = [
                dict(zip(headers, row))
                for row in worksheet.iter_rows(min_row=2, values_only=True)
            ]
            self.assertIn("Общая задолженность", headers)
            self.assertNotIn("ФИО (заёмщика)", headers)
            self.assertEqual([row["Общая задолженность"] for row in rows], [300, 300])
            workbook.close()

    def test_debt_balance_checks_only_dbz_named_in_purpose(self) -> None:
        with TemporaryDirectory() as directory:
            root = Path(directory)
            statement = root / "statement.xlsx"
            reference = root / "reference.xlsx"
            output = root / "output"

            statement_book = Workbook()
            statement_sheet = statement_book.active
            statement_sheet.append(["Дата", "Отправитель", "Кредит", "Назначение платежа"])
            statement_sheet.append(
                ["15.07.2026", "Плательщик", 250, "Оплата ИИН 900101300001"]
            )
            statement_sheet.append(
                [
                    "15.07.2026",
                    "Плательщик",
                    150,
                    "Оплата ИИН 900101300001 ДБЗ: DBZ-1",
                ]
            )
            statement_sheet.append(
                ["15.07.2026", "Плательщик", 150, "Оплата ИИН 900101300001"]
            )
            statement_sheet.append(
                [
                    "15.07.2026",
                    "Плательщик",
                    150,
                    "Оплата ИИН 900101300001 ДБЗ: DBZ-1 DBZ-2",
                ]
            )
            statement_sheet.append(
                [
                    "15.07.2026",
                    "Плательщик",
                    10,
                    "Оплата ИИН 900101300001 ДБЗ: DBZ-4",
                ]
            )
            statement_book.save(statement)

            reference_book = Workbook()
            reference_sheet = reference_book.active
            reference_sheet.append(
                ["Взыскатель", "ДБЗ", "ИИН", "ФИО", "Остаток долга"]
            )
            reference_sheet.append(
                ["Компания", "DBZ-1", "900101300001", "Иванов Иван", 100]
            )
            reference_sheet.append(
                ["Компания", "DBZ-2", "900101300001", "Иванов Иван", 200]
            )
            reference_sheet.append(
                ["Компания", "DBZ-3", "900101300001", "Иванов Иван", 1000]
            )
            reference_sheet.append(
                ["Компания", "DBZ-4", "900101300001", "Иванов Иван", None]
            )
            reference_book.save(reference)

            result = NewBankStatementProcessor(FakeNameExtractor({})).process_many(
                statement_paths=[statement],
                reference_path=reference,
                output_dir=output,
            )

            workbook = load_workbook(result.registry_path, read_only=True, data_only=True)
            payment_headers = [cell.value for cell in workbook["Оплата"][1]]
            payment_rows = [
                dict(zip(payment_headers, row))
                for row in workbook["Оплата"].iter_rows(min_row=2, values_only=True)
            ]
            self.assertEqual(len(payment_rows), 0)

            not_found_headers = [cell.value for cell in workbook["Не найдено"][1]]
            not_found_rows = [
                dict(zip(not_found_headers, row))
                for row in workbook["Не найдено"].iter_rows(min_row=2, values_only=True)
            ]
            self.assertEqual(len(not_found_rows), 5)
            reasons = [row["Причина"].lower() for row in not_found_rows]
            self.assertEqual(sum("переплат" in reason for reason in reasons), 1)
            self.assertEqual(
                sum("два или более дбз" in reason for reason in reasons),
                3,
            )
            self.assertEqual(
                sum(
                    "в назначении платежа найдено два или более дбз" in reason
                    for reason in reasons
                ),
                1,
            )
            self.assertEqual(
                sum("не указан остаток задолженности" in reason for reason in reasons),
                1,
            )
            workbook.close()

    def test_numeric_exclusion_does_not_match_inside_iin(self) -> None:
        self.assertFalse(
            _contains_excluded_keyword("иин 010812600769", "108126")
        )

    def test_numeric_exclusion_matches_standalone_code(self) -> None:
        self.assertTrue(
            _contains_excluded_keyword("назначение платежа 108126", "108126")
        )

    def test_cyrillic_and_latin_dbz_letters_are_equivalent(self) -> None:
        self.assertEqual(
            _normalize_identifier("ЕМ23200-004753436"),
            _normalize_identifier("EM23200-004753436"),
        )

    def test_processor_stops_when_operation_is_cancelled(self) -> None:
        with TemporaryDirectory() as directory:
            with self.assertRaises(ProcessingCancelled):
                NewBankStatementProcessor().process_many(
                    statement_paths=[Path(directory) / "statement.xlsx"],
                    reference_path=Path(directory) / "reference.xlsx",
                    output_dir=Path(directory) / "output",
                    should_cancel=lambda: True,
                )

    def test_name_matching_is_case_insensitive_and_order_independent(self) -> None:
        evidence = match_person_name(
            "Иванов Иван Иванович",
            "иванович ИВАНОВ иВаН",
        )
        self.assertIsNotNone(evidence)
        self.assertEqual(evidence.criterion, "Полное ФИО GLiNER")

    def test_name_matching_accepts_initials_with_or_without_dots(self) -> None:
        for detected in ("Иванов И. И.", "И И ИВАНОВ", "Иванов ИИ"):
            with self.subTest(detected=detected):
                self.assertIsNotNone(
                    match_person_name("Иванов Иван Иванович", detected)
                )

    def test_name_matching_rejects_weak_abbreviations(self) -> None:
        for detected in ("Иванов", "Иванов И", "И И И", "И.И.И."):
            with self.subTest(detected=detected):
                self.assertIsNone(
                    match_person_name("Иванов Иван Иванович", detected)
                )

    def test_multiple_borrowers_in_purpose_are_sent_to_not_found(self) -> None:
        with TemporaryDirectory() as directory:
            root = Path(directory)
            statement = root / "statement.xlsx"
            reference = root / "reference.xlsx"
            output = root / "output"
            purpose = (
                "Должники: Иванов Иван Иванович и Петров Петр Петрович "
                "ИИН 900101300001"
            )

            statement_book = Workbook()
            statement_sheet = statement_book.active
            statement_sheet.append(["Дата", "Отправитель", "Кредит", "Назначение платежа"])
            statement_sheet.append(["15.07.2026", "Плательщик", 1000, purpose])
            statement_book.save(statement)

            reference_book = Workbook()
            reference_sheet = reference_book.active
            reference_sheet.append(["Взыскатель", "ДБЗ", "ИИН", "ФИО"])
            reference_sheet.append(
                ["Компания", "DBZ-1", "900101300001", "Иванов Иван Иванович"]
            )
            reference_sheet.append(
                ["Компания", "DBZ-2", "900101300002", "Петров Петр Петрович"]
            )
            reference_book.save(reference)

            extractor = FakeNameExtractor(
                {
                    purpose: (
                        "Иванов Иван Иванович",
                        "Петров Петр Петрович",
                    )
                }
            )
            result = NewBankStatementProcessor(extractor).process_many(
                statement_paths=[statement],
                reference_path=reference,
                output_dir=output,
            )

            workbook = load_workbook(result.registry_path, read_only=True, data_only=True)
            self.assertEqual(workbook["Оплата"].max_row - 1, 0)
            self.assertEqual(workbook["Не найдено"].max_row - 1, 1)
            reason = workbook["Не найдено"][2][13].value
            self.assertIn("найдено несколько должников", reason.lower())
            workbook.close()

    def test_chsi_counterparty_iin_is_not_used_as_debtor_iin(self) -> None:
        with TemporaryDirectory() as directory:
            root = Path(directory)
            statement = root / "statement.xlsx"
            reference = root / "reference.xlsx"
            output = root / "output"

            statement_book = Workbook()
            statement_sheet = statement_book.active
            statement_sheet.append(
                [
                    "Дата",
                    "Отправитель",
                    "БИН/ИИН контрагента",
                    "Кредит",
                    "Назначение платежа",
                ]
            )
            statement_sheet.append(
                [
                    "15.07.2026",
                    "ЧСИ Талипов Нурмуханович",
                    "661201300130",
                    1000,
                    "Взыскано с должника Хайруллина О.С. ИИН 800908402859",
                ]
            )
            statement_book.save(statement)

            reference_book = Workbook()
            reference_sheet = reference_book.active
            reference_sheet.append(["Взыскатель", "ДБЗ", "ИИН", "ФИО"])
            reference_sheet.append(
                ["Компания", "DBZ-CHSI", "661201300130", "Талипов Нурмухан"]
            )
            reference_sheet.append(
                ["Компания", "DBZ-DEBTOR", "800908402859", "Хайруллина Ольга"]
            )
            reference_book.save(reference)

            result = NewBankStatementProcessor(FakeNameExtractor({})).process_many(
                statement_paths=[statement],
                reference_path=reference,
                output_dir=output,
            )

            workbook = load_workbook(result.registry_path, read_only=True, data_only=True)
            self.assertEqual(workbook["Оплата"].max_row - 1, 1)
            headers = [cell.value for cell in workbook["Оплата"][1]]
            payment = dict(
                zip(
                    headers,
                    next(workbook["Оплата"].iter_rows(min_row=2, values_only=True)),
                )
            )
            self.assertEqual(payment["ИИН"], "800908402859")
            self.assertEqual(payment["ДБЗ"], "DBZ-DEBTOR")
            self.assertEqual(payment["Оплата через"], "ЧСИ")
            workbook.close()

    def test_purpose_iin_has_priority_over_counterparty_iin(self) -> None:
        with TemporaryDirectory() as directory:
            root = Path(directory)
            statement = root / "statement.xlsx"
            reference = root / "reference.xlsx"
            output = root / "output"

            statement_book = Workbook()
            statement_sheet = statement_book.active
            statement_sheet.append(
                [
                    "Дата",
                    "Отправитель",
                    "БИН/ИИН контрагента",
                    "Кредит",
                    "Назначение платежа",
                ]
            )
            statement_sheet.append(
                [
                    "15.07.2026",
                    "Плательщик",
                    "661201300130",
                    1000,
                    "Оплата долга, ИИН должника 800908402859",
                ]
            )
            statement_book.save(statement)

            reference_book = Workbook()
            reference_sheet = reference_book.active
            reference_sheet.append(["Взыскатель", "ДБЗ", "ИИН", "ФИО"])
            reference_sheet.append(
                ["Компания", "DBZ-COUNTERPARTY", "661201300130", "Иванов Иван"]
            )
            reference_sheet.append(
                ["Компания", "DBZ-PURPOSE", "800908402859", "Петров Петр"]
            )
            reference_book.save(reference)

            result = NewBankStatementProcessor(FakeNameExtractor({})).process_many(
                statement_paths=[statement],
                reference_path=reference,
                output_dir=output,
            )

            workbook = load_workbook(result.registry_path, read_only=True, data_only=True)
            headers = [cell.value for cell in workbook["Оплата"][1]]
            payment = dict(
                zip(
                    headers,
                    next(workbook["Оплата"].iter_rows(min_row=2, values_only=True)),
                )
            )
            self.assertEqual(payment["ИИН"], "800908402859")
            self.assertEqual(payment["ДБЗ"], "DBZ-PURPOSE")
            workbook.close()

    def test_name_matching_allows_safe_grammatical_endings(self) -> None:
        reference_name = "Нуриева Сания Ирлановна"
        declined_name = "Нуриевой Сании Ирлановны"
        self.assertIsNone(
            match_person_name(
                reference_name,
                declined_name,
            )
        )
        self.assertIsNotNone(
            match_person_name_nominative(reference_name, declined_name)
        )

    def test_name_matching_rejects_different_names_with_shared_prefix(self) -> None:
        self.assertIsNone(
            match_person_name(
                "Сайдалин Самарбай Муратович",
                "Сайдалин Самат Муратович",
            )
        )
        self.assertIsNone(
            match_person_name_nominative(
                "Сайдалин Самарбай Муратович",
                "Сайдалин Самат Муратович",
            )
        )

    def test_sender_iin_is_rejected_when_sender_name_conflicts(self) -> None:
        with TemporaryDirectory() as directory:
            root = Path(directory)
            statement = root / "statement.xlsx"
            reference = root / "reference.xlsx"
            output = root / "output"
            sender = "Сайдалин Самат Муратович"

            statement_book = Workbook()
            statement_sheet = statement_book.active
            statement_sheet.append(
                [
                    "Дата",
                    "Отправитель",
                    "БИН/ИИН контрагента",
                    "Кредит",
                    "Назначение платежа",
                ]
            )
            statement_sheet.append(
                [
                    "15.07.2026",
                    sender,
                    "820523351247",
                    50000,
                    "Погашение долгосрочного займа",
                ]
            )
            statement_book.save(statement)

            reference_book = Workbook()
            reference_sheet = reference_book.active
            reference_sheet.append(["Взыскатель", "ДБЗ", "ИИН", "ФИО"])
            reference_sheet.append(
                [
                    "Компания",
                    "100000315334/B",
                    "820523351247",
                    "Сайдалин Самарбай Муратович",
                ]
            )
            reference_book.save(reference)

            result = NewBankStatementProcessor(
                FakeNameExtractor({sender: (sender,)})
            ).process_many(
                statement_paths=[statement],
                reference_path=reference,
                output_dir=output,
            )

            workbook = load_workbook(result.registry_path, read_only=True, data_only=True)
            self.assertEqual(workbook["Оплата"].max_row - 1, 0)
            self.assertEqual(workbook["Не найдено"].max_row - 1, 1)
            reason = workbook["Не найдено"][2][13].value
            self.assertIn("ФИО не совпадает", reason)
            workbook.close()

    def test_name_matching_retries_in_nominative_case(self) -> None:
        reference_name = "Петров Пётр Петрович"
        declined_name = "Петровым Петром Петровичем"

        self.assertIsNone(match_person_name(reference_name, declined_name))
        self.assertIsNotNone(
            match_person_name_nominative(reference_name, declined_name)
        )

    def test_name_matching_normalizes_declined_reference_column(self) -> None:
        self.assertIsNotNone(
            match_person_name_nominative(
                "Петровым Петром Петровичем",
                "Петров Пётр Петрович",
            )
        )

    def test_nominative_kazakh_and_feminine_names_are_not_rewritten(self) -> None:
        expected_names = {
            "Купесбаева Айнұр Бикенқызы": (
                "купесбаева",
                "айнұр",
                "бикенқызы",
            ),
            "Жанбетова Сауле Тыныштекбаевна": (
                "жанбетова",
                "сауле",
                "тыныштекбаевна",
            ),
            "СЕЙДІЛДА ЕЛДОС БӨКЕНҰЛЫ": (
                "сейділда",
                "елдос",
                "бөкенұлы",
            ),
        }
        for name, expected in expected_names.items():
            with self.subTest(name=name):
                self.assertEqual(nominative_person_name_tokens(name), expected)

    def test_real_statement_name_is_converted_from_genitive(self) -> None:
        declined_name = "Зленко Ивана Анатольевича"
        self.assertEqual(
            nominative_person_name_tokens(declined_name),
            ("зленко", "иван", "анатольевич"),
        )
        self.assertIsNone(
            match_person_name("Зленко Иван Анатольевич", declined_name)
        )
        self.assertIsNotNone(
            match_person_name_nominative(
                "Зленко Иван Анатольевич",
                declined_name,
            )
        )

    def test_every_statement_row_is_output_or_has_skip_reason(self) -> None:
        with TemporaryDirectory() as directory:
            root = Path(directory)
            statement = root / "KZ001_Компания.xlsx"
            reference = root / "reference.xlsx"
            output = root / "output"

            statement_book = Workbook()
            statement_sheet = statement_book.active
            statement_sheet.append(["Дата", "Отправитель", "Кредит", "Назначение платежа"])
            statement_sheet.append(["15.07.2026", "Плательщик", 1000, "Оплата Иван Иванович Иванов"])
            statement_sheet.append(["15.07.2026", "Плательщик", 2000, "Оплата Иванов"])
            statement_sheet.append(
                ["15.07.2026", "Плательщик", 3000, "Возврат ошибочных элементов"]
            )
            statement_sheet.append(["15.07.2026", "Плательщик", 0, "Нулевая сумма"])
            statement_book.save(statement)

            reference_book = Workbook()
            reference_sheet = reference_book.active
            reference_sheet.append(["Взыскатель", "ДБЗ", "ИИН", "ФИО"])
            reference_sheet.append(["Компания", "DBZ-1", "900101300001", "Иванов Иван Иванович"])
            reference_book.save(reference)

            extractor = FakeNameExtractor(
                {
                    "Оплата Иван Иванович Иванов": ("Иван Иванович Иванов",),
                    "Оплата Иванов": ("Иванов",),
                }
            )
            result = NewBankStatementProcessor(extractor).process_many(
                statement_paths=[statement],
                reference_path=reference,
                output_dir=output,
            )

            workbook = load_workbook(result.registry_path, read_only=True, data_only=True)
            self.assertEqual(workbook["Оплата"].max_row - 1, 1)
            payment_headers = [cell.value for cell in workbook["Оплата"][1]]
            payment_row = dict(
                zip(
                    payment_headers,
                    next(workbook["Оплата"].iter_rows(min_row=2, values_only=True)),
                )
            )
            self.assertIsNone(payment_row["Общая задолженность"])
            self.assertEqual(workbook["Не найдено"].max_row - 1, 1)
            self.assertEqual(workbook["Журнал пропусков"].max_row - 1, 2)
            reasons = [
                row[6]
                for row in workbook["Журнал пропусков"].iter_rows(
                    min_row=2,
                    values_only=True,
                )
            ]
            self.assertTrue(all(reasons))
            workbook.close()


if __name__ == "__main__":
    unittest.main()
