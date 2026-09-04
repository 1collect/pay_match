from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


class ProcessingError(Exception):
    """Raised when an uploaded workbook cannot be processed safely."""


EXCLUDED_KEYWORDS = (
    "перевод собственных средств",
    "перевод между счетами",
    "\u0432\u043e\u0437\u0432\u0440\u0430\u0442 \u043e\u0448\u0438\u0431\u043e\u0447\u043d\u044b\u0445 \u044d\u043b\u0435\u043c\u0435\u043d\u0442\u043e\u0432",
    "возврат",
    "айналымдар / обороты",
    "обороты:",
    "итого обороты",
    "108126",
)

STATEMENT_SUMMARY_KEYWORDS = (
    "\u0438\u0442\u043e\u0433\u043e \u043e\u0431\u043e\u0440\u043e\u0442\u044b",
    "\u0438\u0442\u043e\u0433\u043e \u043d\u0434\u0441",
    "\u0438\u0442\u043e\u0433\u043e \u043e\u0431\u043e\u0440\u043e\u0442\u043e\u0432",
    "\u0430\u0439\u043d\u0430\u043b\u044b\u043c\u0434\u0430\u0440 \u043e\u0431\u043e\u0440\u043e\u0442\u044b",
)

CREDIT_HEADER_MARKERS = (
    "кредит",
    "сумма кредита",
    "сумма по кредиту",
    "поступление",
    "приход",
    "сумма поступления",
    "credit",
    "amount credit",
)

OUTPUT_COLUMNS = [
    "Взыскатель",
    "ДБЗ (договор займа)",
    "ФИО (заёмщика)",
    "ИИН (заёмщика)",
    "Сумма платежа",
    "Дата платежа",
    "Расчетный счет",
    "Компания",
]

NOT_FOUND_COLUMNS = [
    "Дата платежа",
    "Отправитель",
    "Сумма",
    "Назначение платежа",
    "Расчетный счет",
    "Компания",
    "ИИН из выписки",
    "ФИО из выписки",
    "Исходная строка",
    "Причина",
]

COMPANY_SUMMARY_COLUMNS = [
    "\u041a\u043e\u043c\u043f\u0430\u043d\u0438\u044f",
    "\u041e\u0431\u0449\u0430\u044f \u0441\u0443\u043c\u043c\u0430",
]

STATEMENT_ALIASES = {
    "date": [
        "дата",
        "дата валютирования",
        "дата платежа",
        "дата операции",
        "дата проводки",
        "валютирования",
        "value date",
        "date",
    ],
    "sender": [
        "отправитель",
        "плательщик",
        "контрагент",
        "наименование контрагента",
        "наименование отправителя",
        "наименование плательщика",
        "sender",
        "payer",
    ],
    "amount": [
        "сумма кредита",
        "сумма по кредиту",
        "кредит",
        "поступление",
        "приход",
        "сумма поступления",
        "credit",
        "amount credit",
        "сумма",
    ],
    "purpose": [
        "назначение платежа",
        "назначение",
        "детали платежа",
        "описание платежа",
        "описание",
        "payment details",
        "purpose",
    ],
    "account": [
        "расчетный счет",
        "расчётный счет",
        "р/с",
        "счет",
        "счёт",
        "номер счета",
        "iban",
        "account",
    ],
    "company": [
        "компания",
        "организация",
        "получатель",
        "бенефициар",
        "наименование компании",
        "company",
    ],
    "counterparty_iin": [
        "бин/иин контрагента",
        "бин иин контрагента",
        "иин контрагента",
        "бин контрагента",
        "бсн/жсн контрагента",
        "bin/iin counterparty",
    ],
}

REFERENCE_ALIASES = {
    "creditor": ["взыскатель", "кредитор", "creditor"],
    "dbz": ["дбз", "договор займа", "договор", "номер договора", "dbz"],
    "iin": ["иин", "iin"],
    "fio": ["фио", "фио заемщика", "фио заёмщика", "заемщик", "заёмщик", "должник", "fio"],
    "variants": [
        "варианты",
        "варианты фио",
        "ключевые слова",
        "возможные варианты",
        "алиасы",
        "aliases",
    ],
}

MONEY_PATTERN = re.compile(r"[-+]?\d[\d\s\u00a0]*(?:[,.]\d{1,2})?")
IIN_PATTERN = re.compile(r"(?<!\d)(\d{12})(?!\d)")
IDENTIFIER_TRANSLATION = str.maketrans(
    {
        "А": "A",
        "В": "B",
        "Е": "E",
        "К": "K",
        "М": "M",
        "Н": "H",
        "О": "O",
        "Р": "P",
        "С": "C",
        "Т": "T",
        "У": "Y",
        "Х": "X",
    }
)
DETAIL_PATTERN = re.compile(
    r"(?P<fio>[A-ZА-ЯЁ][^;\n\r]{1,160}?)\s*"
    r"(?:ИИН|IIN)\s*[:№#]?\s*"
    r"(?P<iin>\d{12})\s*"
    r"[-–—:,\s]*"
    r"(?P<amount>[-+]?\d[\d\s\u00a0]*(?:[,.]\d{1,2})?)",
    flags=re.IGNORECASE,
)

DETAIL_NAME_STOPWORDS = {
    "назначение",
    "платеж",
    "платежа",
    "платежи",
    "общая",
    "сумма",
    "оплата",
    "оплаты",
    "погашение",
    "погашения",
    "за",
    "от",
    "по",
    "договору",
    "дог",
    "займ",
    "займа",
    "кредит",
    "кредита",
    "ип",
    "тоо",
}


@dataclass
class DetectedColumns:
    date: str | None = None
    sender: str | None = None
    amount: str | None = None
    purpose: str | None = None
    account: str | None = None
    company: str | None = None
    counterparty_iin: str | None = None


@dataclass
class StatementTable:
    df: pd.DataFrame
    columns: DetectedColumns
    header_row: int
    filename_account: str = ""
    filename_company: str = ""
    inferred_account: str = ""
    inferred_company: str = ""
    warnings: list[str] | None = None


@dataclass
class PaymentDetail:
    fio: str
    iin: str
    amount: float | None


@dataclass
class PaymentRecord:
    value_date: date | None
    sender: str
    amount: float | None
    purpose: str
    account: str
    company: str
    source_row: int
    detail_fio: str = ""
    detail_iin: str = ""
    counterparty_iin: str = ""


@dataclass
class ReferenceEntry:
    row_id: int
    creditor: str
    dbz: str
    iin: str
    fio: str
    normalized_dbz: str = ""
    normalized_fio: str = ""
    name_pair: str = ""
    surname: str = ""
    keywords: tuple[str, ...] = ()


@dataclass
class MatchCandidate:
    entry: ReferenceEntry
    criteria: set[str] = field(default_factory=set)
    detected_iin: str = ""
    detected_name: str = ""

    @property
    def score(self) -> int:
        return len(self.criteria)


@dataclass
class MatchResult:
    entry: ReferenceEntry | None
    detected_iin: str = ""
    detected_name: str = ""
    criteria: tuple[str, ...] = ()
    reason: str = "Не найдено совпадение в справочнике"
    source: str = ""


@dataclass
class ReferenceLookup:
    by_iin: dict[str, list[ReferenceEntry]]
    by_dbz: dict[str, list[ReferenceEntry]]
    by_name_token: dict[str, list[ReferenceEntry]]
    rows_count: int

    def find(self, record: PaymentRecord) -> MatchResult:
        purpose_text = "\n".join(
            part
            for part in (
                record.detail_fio,
                record.detail_iin,
                record.counterparty_iin,
                record.purpose,
            )
            if part
        )
        purpose_result = self._find_in_text(record, purpose_text, "Назначение платежа")
        if purpose_result.entry:
            return purpose_result
        if _is_blocking_failed_result(purpose_result):
            return purpose_result

        combined_text = "\n".join(
            part
            for part in (
                record.detail_fio,
                record.detail_iin,
                record.counterparty_iin,
                record.purpose,
                record.sender,
            )
            if part
        )
        combined_result = self._find_in_text(record, combined_text, "Назначение платежа + Отправитель")
        if combined_result.entry:
            return combined_result
        if _is_blocking_failed_result(combined_result):
            return combined_result

        sender_result = self._find_in_text(record, record.sender, "Отправитель")
        if sender_result.entry:
            return sender_result

        return _best_failed_result(purpose_result, combined_result, sender_result)

    def _find_in_text(self, record: PaymentRecord, text: str, source: str) -> MatchResult:
        normalized_text = _normalize_name(text)
        text_tokens = set(_name_tokens(text))
        candidates: dict[int, MatchCandidate] = {}

        for iin in _extract_iins(text):
            for entry in self.by_iin.get(iin, []):
                candidate = _candidate_for(candidates, entry)
                candidate.criteria.add("ИИН")
                candidate.detected_iin = iin

        for dbz_key in _extract_dbz_keys(text):
            for entry in self.by_dbz.get(dbz_key, []):
                _candidate_for(candidates, entry).criteria.add("ДБЗ")

        name_entries: dict[int, ReferenceEntry] = {}
        for token in text_tokens:
            for entry in self.by_name_token.get(token, []):
                name_entries[entry.row_id] = entry

        for entry in name_entries.values():
            name_criterion = _name_match_criterion(entry, normalized_text, text_tokens)
            if not name_criterion:
                continue
            candidate = _candidate_for(candidates, entry)
            candidate.criteria.add(name_criterion)
            candidate.detected_name = _detected_name_for(entry, name_criterion)

        if _is_positive_amount(record.amount):
            for candidate in candidates.values():
                candidate.criteria.add("Сумма по кредиту")

        valid_candidates = [candidate for candidate in candidates.values() if candidate.score >= 2]
        if not valid_candidates:
            best = _top_candidate(candidates.values())
            if _has_multiple_dbz(candidates.values()):
                return MatchResult(
                    entry=None,
                    detected_iin=best.detected_iin if best else "",
                    detected_name=best.detected_name if best else "",
                    criteria=tuple(sorted(best.criteria)) if best else (),
                    reason="По найденному критерию в справочнике найдено несколько ДБЗ",
                    source=source,
                )
            if best:
                return MatchResult(
                    entry=None,
                    detected_iin=best.detected_iin,
                    detected_name=best.detected_name,
                    criteria=tuple(sorted(best.criteria)),
                    reason="Недостаточно критериев для подтверждения заемщика",
                    source=source,
                )
            return MatchResult(
                entry=None,
                reason="Не найдено совпадение в справочнике",
                source=source,
            )

        top_score = max(candidate.score for candidate in valid_candidates)
        top_candidates = [candidate for candidate in valid_candidates if candidate.score == top_score]
        if len(top_candidates) > 1:
            best = top_candidates[0]
            reason = (
                "По найденному критерию в справочнике найдено несколько ДБЗ"
                if _has_multiple_dbz(top_candidates)
                else "Найдено неоднозначное совпадение по нескольким заемщикам"
            )
            return MatchResult(
                entry=None,
                detected_iin=best.detected_iin,
                detected_name=best.detected_name,
                criteria=tuple(sorted(best.criteria)),
                reason=reason,
                source=source,
            )

        best = top_candidates[0]
        ambiguous_criterion = _ambiguous_strong_criterion(valid_candidates)
        if ambiguous_criterion and "ДБЗ" not in best.criteria:
            return MatchResult(
                entry=None,
                detected_iin=best.detected_iin,
                detected_name=best.detected_name,
                criteria=tuple(sorted(best.criteria)),
                reason=f"По критерию «{ambiguous_criterion}» в справочнике найдено несколько ДБЗ",
                source=source,
            )

        return MatchResult(
            entry=best.entry,
            detected_iin=best.detected_iin,
            detected_name=best.detected_name,
            criteria=tuple(sorted(best.criteria)),
            source=source,
        )


@dataclass
class ProcessingResult:
    registry_path: Path
    not_found_path: Path
    total_statement_rows: int
    generated_records: int
    matched_rows: int
    not_found_rows: int
    excluded_rows: int
    empty_amount_rows: int
    split_rows: int
    reference_rows: int
    warnings: list[str]


def _build_company_summary_df(
    registry_rows: list[dict[str, Any]],
    not_found_rows: list[dict[str, Any]],
    *,
    registry_company_column: str,
    registry_amount_column: str,
    not_found_company_column: str,
    not_found_amount_column: str,
) -> pd.DataFrame:
    totals: dict[str, dict[str, Any]] = {}

    def add_rows(rows: list[dict[str, Any]], company_column: str, amount_column: str) -> None:
        for row in rows:
            amount = _parse_amount(row.get(amount_column))
            if amount is None:
                continue
            company = _summary_company_name(row.get(company_column))
            key = company.casefold()
            if key not in totals:
                totals[key] = {
                    "\u041a\u043e\u043c\u043f\u0430\u043d\u0438\u044f": company,
                    "\u041e\u0431\u0449\u0430\u044f \u0441\u0443\u043c\u043c\u0430": 0.0,
                }
            totals[key]["\u041e\u0431\u0449\u0430\u044f \u0441\u0443\u043c\u043c\u0430"] += float(amount)

    add_rows(registry_rows, registry_company_column, registry_amount_column)
    add_rows(not_found_rows, not_found_company_column, not_found_amount_column)

    rows = sorted(totals.values(), key=lambda row: row["\u041a\u043e\u043c\u043f\u0430\u043d\u0438\u044f"].casefold())
    rows.append(
        {
            "\u041a\u043e\u043c\u043f\u0430\u043d\u0438\u044f": "\u0418\u0442\u043e\u0433\u043e",
            "\u041e\u0431\u0449\u0430\u044f \u0441\u0443\u043c\u043c\u0430": sum(
                float(row["\u041e\u0431\u0449\u0430\u044f \u0441\u0443\u043c\u043c\u0430"] or 0)
                for row in rows
            ),
        }
    )
    return pd.DataFrame(rows, columns=COMPANY_SUMMARY_COLUMNS)


def _summary_company_name(value: Any) -> str:
    company = re.sub(r"\s+", " ", _clean_cell(value)).strip()
    return company or "\u0411\u0435\u0437 \u043a\u043e\u043c\u043f\u0430\u043d\u0438\u0438"


def _candidate_for(
    candidates: dict[int, MatchCandidate],
    entry: ReferenceEntry,
) -> MatchCandidate:
    if entry.row_id not in candidates:
        candidates[entry.row_id] = MatchCandidate(entry=entry)
    return candidates[entry.row_id]


def _top_candidate(candidates: Any) -> MatchCandidate | None:
    candidate_list = list(candidates)
    if not candidate_list:
        return None
    return max(candidate_list, key=lambda candidate: candidate.score)


def _has_multiple_dbz(candidates: Any) -> bool:
    dbz_values = {
        _normalize_identifier(candidate.entry.dbz)
        for candidate in candidates
        if _normalize_identifier(candidate.entry.dbz)
    }
    return len(dbz_values) > 1


def _ambiguous_strong_criterion(candidates: Any) -> str:
    strong_criteria = ("ИИН", "ФИО", "Фамилия имя", "Ключевое слово")
    candidate_list = list(candidates)
    for criterion in strong_criteria:
        dbz_values = {
            _normalize_identifier(candidate.entry.dbz)
            for candidate in candidate_list
            if criterion in candidate.criteria and _normalize_identifier(candidate.entry.dbz)
        }
        if len(dbz_values) > 1:
            return criterion
    return ""


def _best_failed_result(*results: MatchResult) -> MatchResult:
    best = max(results, key=lambda result: len(result.criteria), default=None)
    if best and (best.detected_iin or best.detected_name or best.criteria):
        return best
    return MatchResult(entry=None)


def _is_blocking_failed_result(match: MatchResult) -> bool:
    if match.entry:
        return False
    lowered_reason = match.reason.lower()
    return "несколько дбз" in lowered_reason or "неоднознач" in lowered_reason


def _format_not_found_reason(match: MatchResult) -> str:
    parts = [match.reason]
    if match.source:
        parts.append(f"поле: {match.source}")
    if match.criteria:
        parts.append("критерии: " + ", ".join(match.criteria))
    return "; ".join(parts)


def _entry_lookup_tokens(entry: ReferenceEntry) -> set[str]:
    tokens: set[str] = set()
    for phrase in (entry.normalized_fio, entry.name_pair, entry.surname, *entry.keywords):
        if not phrase:
            continue
        phrase_tokens = phrase.split()
        if phrase_tokens and _is_useful_name_token(phrase_tokens[0]):
            tokens.add(phrase_tokens[0])
    return tokens


def _name_match_criterion(
    entry: ReferenceEntry,
    normalized_text: str,
    text_tokens: set[str],
) -> str:
    if entry.normalized_fio and _phrase_in_text(entry.normalized_fio, normalized_text):
        return "ФИО"
    if entry.name_pair and _phrase_in_text(entry.name_pair, normalized_text):
        return "Фамилия имя"
    for keyword in entry.keywords:
        if _phrase_in_text(keyword, normalized_text):
            return "Ключевое слово"
    if entry.surname and entry.surname in text_tokens:
        return "Фамилия"
    return ""


def _detected_name_for(entry: ReferenceEntry, criterion: str) -> str:
    if criterion in {"ФИО", "Фамилия имя"}:
        return _clean_output_fio(entry.fio)
    if criterion == "Фамилия":
        return entry.surname
    return _clean_output_fio(entry.fio)


def _clean_output_fio(value: Any) -> str:
    text = _clean_cell(value)
    text = re.sub(r"(?<![A-Za-zА-Яа-яЁё])ЧСИ(?![A-Za-zА-Яа-яЁё])", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", " ", text).strip(" ,.;:-")
    return text


def _phrase_in_text(phrase: str, normalized_text: str) -> bool:
    if not phrase or not normalized_text:
        return False
    return f" {phrase} " in f" {normalized_text} "


def _is_useful_name_phrase(value: str) -> bool:
    tokens = value.split()
    return bool(tokens and any(_is_useful_name_token(token) for token in tokens))


def _is_useful_name_token(value: str) -> bool:
    return bool(value and len(value) >= 2 and not value.isdigit())


def _name_tokens(value: Any) -> list[str]:
    return [
        token
        for token in _normalize_name(value).split()
        if _is_useful_name_token(token)
    ]


def _extract_iins(value: str) -> list[str]:
    return list(dict.fromkeys(IIN_PATTERN.findall(value or "")))


def _extract_dbz_keys(value: str) -> list[str]:
    text = value or ""
    keys: list[str] = []
    label_pattern = re.compile(
        r"(?:дбз|договор(?:\s+банковского)?\s+займа|номер\s+договора|договор)\s*"
        r"[:№#-]?\s*([A-Za-zА-Яа-яЁё0-9][A-Za-zА-Яа-яЁё0-9/_-]{2,60})",
        flags=re.IGNORECASE,
    )
    for match in label_pattern.finditer(text):
        key = _normalize_identifier(match.group(1))
        if key:
            keys.append(key)

        # A document number is occasionally typed with one or two accidental
        # spaces inside it. Add joined variants of the chunks immediately after
        # the first one; only a variant equal to a DBZ from the reference book
        # will be used by the matcher.
        joined_value = match.group(1)
        cursor = match.end(1)
        spaces_used = 0
        while spaces_used < 2:
            continuation = re.match(
                r"([ \t\u00a0]{1,2})([A-Za-zА-Яа-яЁё0-9][A-Za-zА-Яа-яЁё0-9/_-]{0,60})",
                text[cursor:],
            )
            if not continuation:
                break
            spaces_used += len(continuation.group(1))
            if spaces_used > 2:
                break
            joined_value += continuation.group(2)
            joined_key = _normalize_identifier(joined_value)
            if joined_key:
                keys.append(joined_key)
            cursor += continuation.end()

    for token in re.findall(r"[A-Za-zА-Яа-яЁё0-9][A-Za-zА-Яа-яЁё0-9/_-]{4,60}", text):
        key = _normalize_identifier(token)
        if key and not key.isdigit():
            keys.append(key)

    return list(dict.fromkeys(keys))


def _normalize_identifier(value: Any) -> str:
    text = _clean_cell(value).upper().replace("Ё", "Е").translate(IDENTIFIER_TRANSLATION)
    return re.sub(r"[^0-9A-ZА-Я]+", "", text)


def _is_positive_amount(value: float | None) -> bool:
    return value is not None and value > 0


def _is_amount_only_row(row: pd.Series, amount_column: str | None) -> bool:
    if not amount_column:
        return False
    amount = _parse_amount(row.get(amount_column))
    if not _is_positive_amount(amount):
        return False
    for column, value in row.items():
        if column == amount_column:
            continue
        if _clean_cell(value):
            return False
    return True


def _is_statement_summary_row(row: pd.Series, amount_column: str | None) -> bool:
    if not amount_column:
        return False
    amount = _parse_amount(row.get(amount_column))
    if not _is_positive_amount(amount):
        return False

    normalized_row_text = _normalize_name(
        " ".join(_clean_cell(value) for value in row.tolist())
    )
    return any(
        _phrase_in_text(_normalize_name(keyword), normalized_row_text)
        for keyword in STATEMENT_SUMMARY_KEYWORDS
    )


def _read_excel_raw(path: Path, label: str) -> pd.DataFrame:
    try:
        with pd.ExcelFile(path) as workbook:
            last_empty_shape = (0, 0)
            for sheet_name in workbook.sheet_names:
                df = pd.read_excel(workbook, sheet_name=sheet_name, header=None, dtype=object)
                if not df.dropna(how="all").empty:
                    return df
                last_empty_shape = df.shape
    except Exception as exc:
        raise ProcessingError(f"Не удалось прочитать {label}: {path.name}") from exc

    raise ProcessingError(
        f"В файле {path.name} не найдено данных. Проверены листы Excel, "
        f"последний размер таблицы: {last_empty_shape[0]}x{last_empty_shape[1]}."
    )


def _detect_header_row(raw: pd.DataFrame, aliases: dict[str, list[str]]) -> int:
    best_row = 0
    best_score = -1
    max_rows = min(len(raw), 80)

    for row_index in range(max_rows):
        normalized_cells = [_normalize_header(value) for value in raw.iloc[row_index].tolist()]
        score = 0
        for key, key_aliases in aliases.items():
            if any(_cell_matches_aliases(cell, key_aliases) for cell in normalized_cells):
                score += 3 if key in {"date", "amount", "purpose", "iin", "fio"} else 1
        if score > best_score:
            best_score = score
            best_row = row_index

    return best_row


def _detect_statement_header_row(raw: pd.DataFrame) -> int:
    best_row: int | None = None
    best_score = -1
    max_rows = min(len(raw), 120)

    for row_index in range(max_rows):
        normalized_cells = [_normalize_header(value) for value in raw.iloc[row_index].tolist()]
        if not any(_is_credit_header_cell(cell) for cell in normalized_cells):
            continue

        score = 10
        for key, key_aliases in STATEMENT_ALIASES.items():
            if any(_cell_matches_aliases(cell, key_aliases) for cell in normalized_cells):
                score += 3 if key in {"date", "amount", "purpose"} else 1
        if any("дебет" in cell or "debit" in cell for cell in normalized_cells):
            score += 1
        if any(cell in {"n", "no", "номер"} or cell.startswith("№") for cell in normalized_cells):
            score += 1

        if score > best_score:
            best_score = score
            best_row = row_index

    if best_row is None:
        raise ProcessingError(
            "В выписке не найдена строка заголовков с колонкой «Кредит». "
            "Проверьте, что файл содержит табличную выписку, а не пустой лист или сводку."
        )

    return best_row


def _is_credit_header_cell(cell: str) -> bool:
    if not cell or _should_skip_candidate("amount", cell):
        return False
    if any(marker == cell for marker in CREDIT_HEADER_MARKERS):
        return True
    if any(marker in cell for marker in CREDIT_HEADER_MARKERS if marker != "кредит"):
        return True
    return bool(re.search(r"(^|\s)кредит(а)?(\s|$)", cell))


def _build_headers(row: pd.Series) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}

    for column_index, value in enumerate(row.tolist(), start=1):
        header = _clean_cell(value) or f"Колонка {column_index}"
        count = seen.get(header, 0)
        seen[header] = count + 1
        if count:
            header = f"{header} {count + 1}"
        headers.append(header)

    return headers


def _find_column(
    headers: list[str],
    key: str,
    aliases: dict[str, list[str]],
) -> str | None:
    normalized_headers = [(header, _normalize_header(header)) for header in headers]

    for header, normalized_header in normalized_headers:
        if _should_skip_candidate(key, normalized_header):
            continue
        if normalized_header in aliases[key]:
            return header

    for header, normalized_header in normalized_headers:
        if _should_skip_candidate(key, normalized_header):
            continue
        if _cell_matches_aliases(normalized_header, aliases[key]):
            return header

    return None


def _should_skip_candidate(key: str, normalized_header: str) -> bool:
    if key == "amount":
        return any(
            word in normalized_header
            for word in ("дебет", "расход", "списание", "исход", "debit")
        )
    if key == "account":
        return "корреспондент" in normalized_header
    return False


def _cell_matches_aliases(cell: str, aliases: list[str]) -> bool:
    return any(alias == cell or alias in cell for alias in aliases)


def _normalize_header(value: Any) -> str:
    text = _clean_cell(value).lower().replace("ё", "е")
    text = re.sub(r"[^0-9a-zа-я/]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _matrix_to_text(df: pd.DataFrame) -> str:
    parts: list[str] = []
    for value in df.to_numpy().flatten().tolist():
        text = _clean_cell(value)
        if text:
            parts.append(text)
    return "\n".join(parts)


def _extract_filename_account(path: Path) -> str:
    return _infer_account(path.stem)


def _extract_filename_company(path: Path) -> str:
    stem = path.stem
    account = _extract_filename_account(path)
    company = stem

    if account:
        company = re.sub(re.escape(account), " ", company, flags=re.IGNORECASE)

    company = re.sub(r"(?<!\d)\d{4}-\d{2}-\d{2}(?!\d)", " ", company)
    company = re.sub(r"(?<!\d)\d{8}-\d{4,6}(?!\d)", " ", company)
    company = re.sub(r"(?<![0-9A-Z])utf-?8(?![0-9A-Z])", " ", company, flags=re.IGNORECASE)
    company = re.sub(r"[_\-]+", " ", company)
    company = re.sub(r"\s+", " ", company).strip(" ._-")

    return company or stem


def _infer_company(text: str) -> str:
    match = re.search(
        r"(?:компания|организация|получатель|бенефициар)\s*[:\-]\s*([^\n;]{3,120})",
        text,
        flags=re.IGNORECASE,
    )
    return _clean_cell(match.group(1)) if match else ""


def _infer_account(text: str) -> str:
    iban_match = re.search(r"(?<![0-9A-Z])KZ[0-9A-Z]{18}(?![0-9A-Z])", text, flags=re.IGNORECASE)
    if iban_match:
        return iban_match.group(0).upper()

    account_match = re.search(
        r"(?:р/?с|расчетный счет|расчётный счет|счет|счёт)\s*[:№#-]?\s*([A-ZА-Я0-9]{10,30})",
        text,
        flags=re.IGNORECASE,
    )
    return account_match.group(1).strip() if account_match else ""


def _is_excluded(purpose: str, sender: str = "") -> bool:
    lowered = purpose.lower()
    if "\u0432\u043e\u0437\u0432\u0440\u0430\u0442 \u043e\u0448\u0438\u0431\u043e\u0447\u043d\u044b\u0445 \u044d\u043b\u0435\u043c\u0435\u043d\u0442\u043e\u0432" in lowered:
        return True
    if "\u043f\u0435\u0440\u0435\u0432\u043e\u0434 \u0441\u043e\u0431\u0441\u0442\u0432\u0435\u043d\u043d\u044b\u0445 \u0441\u0440\u0435\u0434\u0441\u0442\u0432" in lowered and _is_chsi_sender(sender):
        return False
    return any(_contains_excluded_keyword(lowered, keyword) for keyword in EXCLUDED_KEYWORDS)


def _contains_excluded_keyword(text: str, keyword: str) -> bool:
    if keyword.isdigit():
        return bool(re.search(rf"(?<!\d){re.escape(keyword)}(?!\d)", text))
    return keyword in text


def _is_chsi_sender(sender: str) -> bool:
    normalized = _normalize_name(sender)
    return (
        "\u0447\u0441\u0438" in normalized.split()
        or "\u0447\u0430\u0441\u0442\u043d\u044b\u0439 \u0441\u0443\u0434\u0435\u0431\u043d\u044b\u0439 \u0438\u0441\u043f\u043e\u043b\u043d\u0438\u0442\u0435\u043b\u044c" in normalized
    )


def _parse_payment_details(purpose: str) -> list[PaymentDetail]:
    if not purpose:
        return []

    details: list[PaymentDetail] = []
    seen: set[tuple[str, str, float | None]] = set()
    normalized_text = purpose.replace("\r\n", "\n").replace("\r", "\n")

    for match in DETAIL_PATTERN.finditer(normalized_text):
        fio = _cleanup_detail_fio(match.group("fio"))
        iin = _normalize_iin(match.group("iin"))
        amount = _parse_amount(match.group("amount"))
        if not iin:
            continue

        key = (fio, iin, amount)
        if key in seen:
            continue
        seen.add(key)
        details.append(PaymentDetail(fio=fio, iin=iin, amount=amount))

    return details


def _cleanup_detail_fio(value: str) -> str:
    text = re.sub(r"[^A-Za-zА-Яа-яЁё.\-\s]+", " ", _clean_cell(value))
    tokens = []
    for token in text.split():
        normalized_token = _normalize_name(token)
        if normalized_token in DETAIL_NAME_STOPWORDS:
            continue
        tokens.append(token)

    if len(tokens) > 4:
        tokens = tokens[-4:]
    return " ".join(tokens).strip(" .,-")


def _parse_amount(value: Any) -> float | None:
    if _is_empty(value):
        return None
    if isinstance(value, int | float):
        if pd.isna(value):
            return None
        return round(float(value), 2)

    text = _clean_cell(value)
    matches = MONEY_PATTERN.findall(text)
    if not matches:
        return None

    number = matches[-1].replace("\u00a0", " ").replace(" ", "").replace(",", ".")
    try:
        return round(float(number), 2)
    except ValueError:
        return None


def _parse_date(value: Any) -> date | None:
    if _is_empty(value):
        return None
    text = _clean_cell(value)
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        parsed = pd.to_datetime(text, format="%Y-%m-%d", errors="coerce")
        if pd.isna(parsed):
            return None
        return parsed.date()
    if re.match(r"^\d{4}-\d{2}-\d{2}(?:T|\s)", text):
        parsed = pd.to_datetime(text, dayfirst=False, errors="coerce")
        if pd.isna(parsed):
            return None
        return parsed.date()
    parsed = pd.to_datetime(value, dayfirst=True, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def _normalize_iin(value: Any) -> str:
    if _is_empty(value):
        return ""
    text = _clean_cell(value)
    digits = re.sub(r"\D", "", text)
    if len(digits) == 12:
        return digits
    match = IIN_PATTERN.search(text)
    return match.group(1) if match else ""


def _extract_first_iin(value: str) -> str:
    match = IIN_PATTERN.search(value or "")
    return match.group(1) if match else ""


def _normalize_name(value: Any) -> str:
    text = _clean_cell(value).lower().replace("ё", "е")
    text = re.sub(r"[^\w]+", " ", text, flags=re.UNICODE)
    text = text.replace("_", " ")
    return re.sub(r"\s+", " ", text).strip()


def _split_variants(value: Any) -> list[str]:
    text = _clean_cell(value)
    if not text:
        return []
    return [part.strip() for part in re.split(r"[;\n|,]+", text) if part.strip()]


def _clean_cell(value: Any) -> str:
    if _is_empty(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value).replace("\u00a0", " ")).strip()


def _is_empty(value: Any) -> bool:
    if value is None:
        return True
    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        return False


def _format_workbook(workbook) -> None:
    header_fill = PatternFill("solid", fgColor="E8EEF8")
    header_font = Font(bold=True)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            header = str(worksheet.cell(1, column_cells[0].column).value or "")
            is_date_column = "Дата" in header
            is_amount_column = (
                "Сумма" in header
                or "задолженность" in header.lower()
            )
            for cell in column_cells:
                value = cell.value
                if value is None:
                    continue
                max_length = max(max_length, len(str(value)))
                if cell.row > 1 and is_date_column:
                    cell.number_format = "DD.MM.YYYY"
                elif cell.row > 1 and is_amount_column:
                    cell.number_format = '#,##0.00'

            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)
