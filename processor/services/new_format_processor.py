from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable

import pandas as pd
from openpyxl import load_workbook

from .excel_processor import (
    ProcessingError,
    ProcessingResult,
    _build_company_summary_df,
    _build_headers,
    _clean_cell,
    _clean_output_fio,
    _contains_excluded_keyword,
    _extract_dbz_keys,
    _extract_filename_account,
    _extract_filename_company,
    _extract_iins,
    _find_column,
    _format_workbook,
    _is_positive_amount,
    _is_credit_header_cell,
    _is_amount_only_row,
    _is_statement_summary_row,
    _name_tokens,
    _normalize_identifier,
    _normalize_iin,
    _normalize_name,
    _parse_amount,
    _parse_date,
    _phrase_in_text,
    _read_excel_raw,
    _split_variants,
)
from .name_matching import match_person_name, name_lookup_keys
from .ner_service import PersonNameExtractor


NEW_OUTPUT_COLUMNS = [
    "ДБЗ",
    "ИИН",
    "ФИО (заёмщика)",
    "Дата",
    "Отправитель",
    "Сумма",
    "Назначение платежа",
    "Оплата через",
    "Номерсчета",
    "компания",
    "Взыскатель",
]

NEW_NOT_FOUND_COLUMNS = NEW_OUTPUT_COLUMNS + [
    "ФИО из выписки",
    "Исходная строка",
    "Причина",
]

NEW_RECONCILIATION_COLUMNS = [
    "Тип расшифровки",
    "Сумма выписки",
    "Сумма расшифровки",
    "Разница",
    "Файлы расшифровки",
    "Дата",
    "Отправитель",
    "Назначение платежа",
    "Причина",
]

SKIPPED_ROW_COLUMNS = [
    "Файл выписки",
    "Исходная строка",
    "Дата",
    "Отправитель",
    "Сумма",
    "Назначение платежа",
    "Причина пропуска",
]

NEW_STATEMENT_ALIASES = {
    "date": [
        "дата",
        "дата валютирования",
        "дата документа",
        "дата операции",
        "value date",
        "date",
    ],
    "amount": [
        "сумма",
        "сумма по кредиту",
        "сумма кредита",
        "кредит",
        "credit",
        "amount",
    ],
    "sender": [
        "отправитель",
        "наименование контрагента",
        "контрагент",
        "плательщик",
        "sender",
        "payer",
    ],
    "counterparty": [
        "наименование контрагента",
        "контрагент",
        "отправитель",
    ],
    "purpose": [
        "назначение платежа",
        "назначение",
        "детали платежа",
        "описание платежа",
        "purpose",
    ],
    "payment_method": [
        "оплата через",
        "способ оплаты",
        "канал оплаты",
        "payment method",
    ],
    "account": [
        "номерсчета",
        "номер счета",
        "номер счета клиента",
        "расчетный счет",
        "расчётный счет",
        "iban",
        "account",
    ],
    "iin": [
        "бин/иин контрагента",
        "бин иин контрагента",
        "иин контрагента",
        "бин контрагента",
        "иин",
        "iin",
    ],
}

NEW_REFERENCE_ALIASES = {
    "creditor": ["взыскатель", "кредитор", "creditor"],
    "dbz": ["дбз", "договор займа", "договор", "номер договора", "dbz"],
    "iin": ["иин", "iin"],
    "fio": ["фио", "фио заемщика", "фио заёмщика", "заемщик", "заёмщик", "должник", "fio"],
    "debt_balance": ["остаток долга", "остаток задолженности", "debt balance"],
    "variants": [
        "варианты",
        "варианты фио",
        "ключевые слова",
        "возможные варианты",
        "алиасы",
        "aliases",
    ],
}

NEW_EXCLUDED_PURPOSE_KEYWORDS = (
    "перевод собственных средств",
    "перевод между счетами",
    "\u0432\u043e\u0437\u0432\u0440\u0430\u0442 \u043e\u0448\u0438\u0431\u043e\u0447\u043d\u044b\u0445 \u044d\u043b\u0435\u043c\u0435\u043d\u0442\u043e\u0432",
    "возврат",
    "итого ндс",
    "айналымдар / обороты",
    "обороты:",
    "итого обороты",
    "108126",
)

NEW_PURPOSE_NOT_FOUND_KEYWORDS = (
    "9005981",
)

NEW_SENDER_NOT_FOUND_KEYWORDS = (
    "astana-plat",
    "astana plat",
    "казпочта",
)

DETAIL_TYPE_LABELS = {
    "astana": "Astana-Plat",
    "kazpost": "Казпочта",
    "halyk": "Народный",
    "eurasian": "ЕВРАЗИЙСКИЙ БАНК",
}


class ProcessingCancelled(Exception):
    """Raised when a running workbook operation is cancelled by its owner."""

DETAIL_ALIASES = {
    "astana": {
        "amount": [
            "сумма",
            "сумма платежа",
            "сумма оплаты",
            "оплата",
            "оплачено",
            "amount",
        ],
        "iin": [
            "договор/лицевой счет",
            "договор лицевой счет",
            "лицевой счет",
            "договор",
            "номер договора",
            "account",
            "contract",
        ],
        "fio": [
            "фио",
            "фио плательщика",
            "плательщик",
            "клиент",
        ],
    },
    "kazpost": {
        "amount": [
            "сумма",
            "сумма платежа",
            "сумма оплаты",
            "оплата",
            "оплачено",
            "amount",
        ],
        "iin": [
            "иин/бин плательщика",
            "иин бин плательщика",
            "бин/иин плательщика",
            "бин иин плательщика",
            "иин плательщика",
            "бин плательщика",
            "иин",
            "бин/иин",
            "iin",
        ],
        "fio": [
            "фио плательщика",
            "фио",
            "плательщик",
            "payer",
        ],
    },
    "halyk": {
        "amount": [
            "сумма",
            "сумма платежа",
            "сумма оплаты",
            "оплата",
            "оплачено",
            "amount",
        ],
        "iin": [
            "иин",
            "иин заемщика",
            "иин заёмщика",
            "iin",
        ],
        "fio": [
            "фио заемщика",
            "фио заёмщика",
            "фио",
            "заемщик",
            "заёмщик",
        ],
    },
    "eurasian": {
        "amount": [
            "сумма",
            "сумма платежа",
            "сумма оплаты",
            "оплата",
            "оплачено",
            "amount",
        ],
        "iin": [
            "иин/бин плательщика",
            "иин бин плательщика",
            "бин/иин плательщика",
            "бин иин плательщика",
            "иин плательщика",
            "бин плательщика",
            "иин",
            "бин/иин",
            "iin",
        ],
        "fio": [
            "фио плательщика",
            "фио",
            "плательщик",
            "payer",
        ],
    },
}

MONEY_TOLERANCE = 0.01
DETAIL_TYPES_IGNORE_OVERPAYMENT = {"eurasian"}
STATEMENT_RECONCILIATION_ERROR_PURPOSE_KEYWORDS = (
    "оказании услуг по взысканию",
)

COMPANY_ALIASES = {
    "pkb": (
        "ка пкб",
        "первое коллекторское бюро",
        "коллекторское агентство первое коллекторское бюро",
    ),
    "ic": (
        "ка ic",
        "ка айси",
        "коллекторское агентство ic",
        "коллекторское агентство айси",
    ),
    "fc": (
        "ка fc",
        "ка фс",
        "коллекторское агентство fc",
        "коллекторское агентство f collect",
        "коллекторскоеагентство f collect",
        "f collect",
        "fcollect",
    ),
    "alem": (
        "сфк alem",
        "alem finance",
        "алем finance",
        "алем финанс",
    ),
    "almaty_finance": (
        "сфк almaty finance",
        "almaty finance",
        "алматы finance",
        "алматы финанс",
    ),
}

CHSI_PAYMENT_KEYWORDS = (
    "чси",
    "частный судебный исполнитель",
)

WITHHOLDING_PAYMENT_KEYWORDS = (
    "тоо",
    "ргу",
    "кгп",
    "пхв",
    "кгкп",
    "ип",
    "коммунальное государственное учреждение",
    "к/х",
    "коммуналдық",
    "школа",
    "пк",
    "гу",
    "ргкп",
    "сф",
    "гккп",
    "индивидуальный предприниматель",
    "фргп",
    "некоммерческое акционерное общество",
    "чу",
)


@dataclass
class NewColumns:
    date: str | None = None
    amount: str | None = None
    sender: str | None = None
    counterparty: str | None = None
    purpose: str | None = None
    payment_method: str | None = None
    account: str | None = None
    iin: str | None = None


@dataclass
class NewStatementTable:
    df: pd.DataFrame
    columns: NewColumns
    header_row: int
    filename_account: str
    filename_company: str
    source_file: str


@dataclass
class NewPaymentRecord:
    date: Any
    sender: str
    amount: float
    purpose: str
    payment_method: str
    account: str
    company: str
    source_row: int
    counterparty_iin: str = ""
    source_file: str = ""
    purpose_names: tuple[str, ...] = ()
    sender_names: tuple[str, ...] = ()
    ner_reason: str = ""


@dataclass
class NewReferenceEntry:
    row_id: int
    creditor: str
    dbz: str
    iin: str
    fio: str
    normalized_dbz: str
    normalized_fio: str
    name_pair: str
    surname: str
    debt_balance: float | None = None
    keywords: tuple[str, ...] = ()


@dataclass
class NewCandidate:
    entry: NewReferenceEntry
    criteria: set[str] = field(default_factory=set)
    detected_iin: str = ""
    detected_name: str = ""
    name_strength: int = 0

    @property
    def score(self) -> int:
        return len(self.criteria)


@dataclass
class NewMatch:
    entry: NewReferenceEntry | None
    criteria: tuple[str, ...] = ()
    detected_iin: str = ""
    detected_name: str = ""
    reason: str = "Не найдено совпадение в справочнике"
    source: str = ""


@dataclass
class NewReferenceBook:
    entries: list[NewReferenceEntry]
    by_iin: dict[str, list[NewReferenceEntry]]
    by_dbz: dict[str, list[NewReferenceEntry]]
    by_name_token: dict[str, list[NewReferenceEntry]]
    dbz_by_iin: dict[str, set[str]]
    dbz_by_fio: dict[str, set[str]]

    @property
    def rows_count(self) -> int:
        return len(self.entries)


@dataclass
class DetailColumns:
    amount: str | None = None
    iin: str | None = None
    fio: str | None = None


@dataclass
class DetailPaymentRow:
    detail_type: str
    source_file: str
    source_row: int
    iin: str
    fio: str
    amount: float


@dataclass
class DetailBatch:
    detail_type: str
    company_key: str = ""
    company_name: str = ""
    rows: list[DetailPaymentRow] = field(default_factory=list)
    files: list[str] = field(default_factory=list)

    @property
    def total_amount(self) -> float:
        return _round_money(sum(row.amount for row in self.rows))


def _check_cancelled(should_cancel: Callable[[], bool] | None) -> None:
    if should_cancel and should_cancel():
        raise ProcessingCancelled("Операция отменена")


def _report_progress(
    on_progress: Callable[[int, str], None] | None,
    progress: int,
    message: str,
) -> None:
    if on_progress:
        on_progress(progress, message)


class NewBankStatementProcessor:
    def __init__(self, name_extractor: PersonNameExtractor | None = None) -> None:
        self.name_extractor = name_extractor or PersonNameExtractor()

    def process_many(
        self,
        statement_paths: list[Path],
        reference_path: Path,
        output_dir: Path,
        detail_paths: list[Path] | None = None,
        should_cancel: Callable[[], bool] | None = None,
        on_progress: Callable[[int, str], None] | None = None,
    ) -> ProcessingResult:
        output_dir.mkdir(parents=True, exist_ok=True)
        if not statement_paths:
            raise ProcessingError("Не выбраны банковские выписки для обработки.")

        _check_cancelled(should_cancel)
        _report_progress(on_progress, 3, "Читаем справочник заемщиков")
        reference = self._read_reference(reference_path, should_cancel)
        _check_cancelled(should_cancel)
        _report_progress(on_progress, 14, "Читаем файлы расшифровок")
        detail_batches, detail_warnings = self._read_detail_batches(
            detail_paths or [],
            should_cancel,
        )
        matched_rows: list[dict[str, Any]] = []
        not_found_rows: list[dict[str, Any]] = []
        reconciliation_errors: list[dict[str, Any]] = []
        skipped_rows: list[dict[str, Any]] = []
        all_records: list[NewPaymentRecord] = []
        total_rows = 0
        generated_records = 0
        excluded_rows = 0
        empty_amount_rows = 0
        split_rows = 0
        warnings: list[str] = list(detail_warnings)

        statement_count = len(statement_paths)
        for statement_index, statement_path in enumerate(statement_paths, start=1):
            _check_cancelled(should_cancel)
            _report_progress(
                on_progress,
                20 + int(25 * (statement_index - 1) / statement_count),
                f"Читаем выписку {statement_index} из {statement_count}",
            )
            table = self._read_statement(statement_path)
            records, skipped_amount_rows, skipped_excluded_rows, statement_skipped_rows = self._build_records(
                table,
                should_cancel,
            )
            all_records.extend(records)
            skipped_rows.extend(statement_skipped_rows)
            total_rows += len(table.df)
            generated_records += len(records)
            empty_amount_rows += skipped_amount_rows
            excluded_rows += skipped_excluded_rows

        if total_rows != generated_records + excluded_rows + empty_amount_rows:
            raise ProcessingError(
                "Нарушен контроль полноты: не каждая строка выписки получила статус обработки."
            )
        if len(skipped_rows) != excluded_rows + empty_amount_rows:
            raise ProcessingError(
                "Нарушен контроль журнала: для одной или нескольких пропущенных строк нет причины."
            )

        _check_cancelled(should_cancel)
        _report_progress(on_progress, 46, "GLiNER 2 извлекает ФИО из платежей")
        self._extract_record_names(all_records, should_cancel)

        detail_records_by_batch: dict[tuple[str, str], list[NewPaymentRecord]] = {}
        normal_records: list[NewPaymentRecord] = []
        _report_progress(on_progress, 48, "Подготавливаем платежи к сопоставлению")
        for record_index, record in enumerate(all_records):
            if record_index % 100 == 0:
                _check_cancelled(should_cancel)
            reconciliation_reason = _statement_reconciliation_error_reason(record)
            if reconciliation_reason:
                reconciliation_errors.append(
                    _statement_reconciliation_error_row(record, reconciliation_reason)
                )
                continue

            detail_type = _detail_type_for_record(record)
            batch_key = _detail_batch_key_for_record(record, detail_type, detail_batches)
            if batch_key:
                detail_records_by_batch.setdefault(batch_key, []).append(record)
            else:
                normal_records.append(record)

        routed_records = (
            len(normal_records)
            + len(reconciliation_errors)
            + sum(len(records) for records in detail_records_by_batch.values())
        )
        if routed_records != len(all_records):
            raise ProcessingError(
                "Нарушен контроль полноты: один или несколько платежей не попали в маршрут обработки."
            )

        record_count = max(len(normal_records), 1)
        for record_index, record in enumerate(normal_records, start=1):
            if record_index % 25 == 0 or record_index == 1:
                _check_cancelled(should_cancel)
                _report_progress(
                    on_progress,
                    52 + int(30 * record_index / record_count),
                    f"Сопоставляем платежи: {record_index} из {len(normal_records)}",
                )
            self._append_statement_record(record, reference, matched_rows, not_found_rows)

        batch_count = max(len(detail_batches), 1)
        for batch_index, (batch_key, batch) in enumerate(detail_batches.items(), start=1):
            _check_cancelled(should_cancel)
            _report_progress(
                on_progress,
                83 + int(8 * batch_index / batch_count),
                "Сверяем расшифровки платежей",
            )
            detail_type = batch.detail_type
            statement_records = detail_records_by_batch.get(batch_key, [])
            if not statement_records:
                company_note = f" ({batch.company_name})" if batch.company_name else ""
                warnings.append(
                    f"Расшифровка {DETAIL_TYPE_LABELS[detail_type]}{company_note} загружена, "
                    "но в выписках не найден соответствующий платеж."
                )
                continue
            split_rows += self._append_detail_batch(
                detail_type=detail_type,
                batch=batch,
                statement_records=statement_records,
                reference=reference,
                matched_rows=matched_rows,
                not_found_rows=not_found_rows,
                reconciliation_errors=reconciliation_errors,
                should_cancel=should_cancel,
            )

        if reconciliation_errors:
            warnings.append(
                f"Ошибки сверки: {len(reconciliation_errors)}. "
                "Они записаны отдельным листом в итоговый Excel."
            )

        _check_cancelled(should_cancel)
        _report_progress(on_progress, 93, "Формируем итоговый Excel")
        registry_path = output_dir / "registry.xlsx"
        not_found_path = output_dir / "not_found.xlsx"
        self._write_registry(
            registry_path,
            matched_rows,
            not_found_rows,
            reconciliation_errors,
            skipped_rows,
        )
        self._write_not_found(not_found_path, not_found_rows)
        _check_cancelled(should_cancel)
        _report_progress(on_progress, 99, "Подготавливаем файл к скачиванию")

        return ProcessingResult(
            registry_path=registry_path,
            not_found_path=not_found_path,
            total_statement_rows=total_rows,
            generated_records=generated_records,
            matched_rows=len(matched_rows),
            not_found_rows=len(not_found_rows),
            excluded_rows=excluded_rows,
            empty_amount_rows=empty_amount_rows,
            split_rows=split_rows,
            reference_rows=reference.rows_count,
            warnings=warnings,
        )

    def _read_statement(self, path: Path) -> NewStatementTable:
        raw = _read_excel_raw(path, "выписку нового формата")
        header_row = _detect_new_header_row(raw)
        headers = _build_headers(raw.iloc[header_row])
        df = raw.iloc[header_row + 1 :].copy()
        df.columns = headers
        df = df.dropna(how="all")

        columns = NewColumns(
            date=_find_column(headers, "date", NEW_STATEMENT_ALIASES),
            amount=_find_column(headers, "amount", NEW_STATEMENT_ALIASES),
            sender=_find_column(headers, "sender", NEW_STATEMENT_ALIASES),
            counterparty=_find_column(headers, "counterparty", NEW_STATEMENT_ALIASES),
            purpose=_find_column(headers, "purpose", NEW_STATEMENT_ALIASES),
            payment_method=_find_column(headers, "payment_method", NEW_STATEMENT_ALIASES),
            account=_find_column(headers, "account", NEW_STATEMENT_ALIASES),
            iin=_find_column(headers, "iin", NEW_STATEMENT_ALIASES),
        )

        missing_required = []
        if not columns.date:
            missing_required.append("Дата")
        if not columns.amount:
            missing_required.append("Сумма")
        if not columns.sender and not columns.counterparty:
            missing_required.append("Отправитель или Наименование контрагента")
        if missing_required:
            raise ProcessingError(
                "В выписке нового формата не найдены обязательные колонки: "
                + ", ".join(missing_required)
            )

        return NewStatementTable(
            df=df,
            columns=columns,
            header_row=header_row,
            filename_account=_extract_filename_account(path),
            filename_company=_extract_filename_company(path),
            source_file=path.name,
        )

    def _extract_record_names(
        self,
        records: list[NewPaymentRecord],
        should_cancel: Callable[[], bool] | None = None,
    ) -> None:
        if not records:
            return
        try:
            purpose_results = self.name_extractor.extract_many(
                [record.purpose for record in records],
                should_cancel,
            )
            sender_results = self.name_extractor.extract_many(
                [record.sender for record in records],
                should_cancel,
            )
        except InterruptedError as exc:
            raise ProcessingCancelled("Операция отменена") from exc
        except Exception as exc:
            raise ProcessingError(
                "GLiNER 2 не смог обработать назначения платежей. "
                "Проверьте доступность модели и повторите операцию."
            ) from exc

        for record, purpose_result, sender_result in zip(
            records,
            purpose_results,
            sender_results,
            strict=True,
        ):
            record.purpose_names = purpose_result.names
            record.sender_names = sender_result.names
            reasons = [
                reason
                for reason in (purpose_result.reason, sender_result.reason)
                if reason
            ]
            record.ner_reason = "; ".join(dict.fromkeys(reasons))

    def _read_reference(
        self,
        path: Path,
        should_cancel: Callable[[], bool] | None = None,
    ) -> NewReferenceBook:
        raw = _read_excel_raw(path, "справочник")
        header_row = _detect_reference_header_row(raw)
        headers = _build_headers(raw.iloc[header_row])
        df = raw.iloc[header_row + 1 :].copy()
        df.columns = headers
        df = df.dropna(how="all")

        creditor_col = _find_column(headers, "creditor", NEW_REFERENCE_ALIASES)
        dbz_col = _find_column(headers, "dbz", NEW_REFERENCE_ALIASES)
        iin_col = _find_column(headers, "iin", NEW_REFERENCE_ALIASES)
        fio_col = _find_column(headers, "fio", NEW_REFERENCE_ALIASES)
        debt_balance_col = _find_column(headers, "debt_balance", NEW_REFERENCE_ALIASES)
        variants_col = _find_column(headers, "variants", NEW_REFERENCE_ALIASES)

        missing_required = []
        if not dbz_col:
            missing_required.append("ДБЗ")
        if not iin_col:
            missing_required.append("ИИН")
        if not fio_col:
            missing_required.append("ФИО")
        if missing_required:
            raise ProcessingError(
                "В справочнике не найдены обязательные колонки: "
                + ", ".join(missing_required)
            )

        entries: list[NewReferenceEntry] = []
        by_iin: dict[str, list[NewReferenceEntry]] = {}
        by_dbz: dict[str, list[NewReferenceEntry]] = {}
        by_name_token: dict[str, list[NewReferenceEntry]] = {}
        dbz_by_iin: dict[str, set[str]] = {}
        dbz_by_fio: dict[str, set[str]] = {}

        for row_index, (_, row) in enumerate(df.iterrows()):
            if row_index % 100 == 0:
                _check_cancelled(should_cancel)
            dbz = _clean_cell(row.get(dbz_col))
            iin = _normalize_iin(row.get(iin_col))
            fio = _clean_cell(row.get(fio_col))
            if not dbz and not iin and not fio:
                continue

            normalized_fio = _normalize_name(fio)
            name_tokens = _name_tokens(fio)
            keyword_values = _split_variants(row.get(variants_col)) if variants_col else []
            keywords = tuple(
                keyword
                for keyword in (_normalize_name(value) for value in keyword_values)
                if keyword
            )
            entry = NewReferenceEntry(
                row_id=len(entries),
                creditor=_clean_cell(row.get(creditor_col)) if creditor_col else "",
                dbz=dbz,
                iin=iin,
                fio=fio,
                normalized_dbz=_normalize_identifier(dbz),
                normalized_fio=normalized_fio,
                name_pair=" ".join(name_tokens[:2]) if len(name_tokens) >= 2 else "",
                surname=name_tokens[0] if name_tokens else "",
                debt_balance=_parse_amount(row.get(debt_balance_col)) if debt_balance_col else None,
                keywords=keywords,
            )
            entries.append(entry)

            if entry.iin:
                by_iin.setdefault(entry.iin, []).append(entry)
                if entry.normalized_dbz:
                    dbz_by_iin.setdefault(entry.iin, set()).add(entry.normalized_dbz)
            if entry.normalized_dbz:
                by_dbz.setdefault(entry.normalized_dbz, []).append(entry)
            if entry.normalized_fio and entry.normalized_dbz:
                dbz_by_fio.setdefault(entry.normalized_fio, set()).add(entry.normalized_dbz)
            for token in _entry_tokens(entry):
                by_name_token.setdefault(token, []).append(entry)

        if not entries:
            raise ProcessingError("В справочнике не найдено ни одной записи для сопоставления.")

        return NewReferenceBook(
            entries=entries,
            by_iin=by_iin,
            by_dbz=by_dbz,
            by_name_token=by_name_token,
            dbz_by_iin=dbz_by_iin,
            dbz_by_fio=dbz_by_fio,
        )

    def _read_detail_batches(
        self,
        detail_paths: list[Path],
        should_cancel: Callable[[], bool] | None = None,
    ) -> tuple[dict[tuple[str, str], DetailBatch], list[str]]:
        batches: dict[tuple[str, str], DetailBatch] = {}
        warnings: list[str] = []

        for detail_path in detail_paths:
            _check_cancelled(should_cancel)
            detail_type = _detail_type_for_path(detail_path)
            if not detail_type:
                warnings.append(
                    f"Файл расшифровки {detail_path.name} пропущен: не определен тип."
                )
                continue

            rows, company_key, company_name = self._read_detail_rows(
                detail_path,
                detail_type,
                should_cancel,
            )
            batch_key = (detail_type, company_key)
            batch = batches.setdefault(
                batch_key,
                DetailBatch(
                    detail_type=detail_type,
                    company_key=company_key,
                    company_name=company_name,
                ),
            )
            if company_name and not batch.company_name:
                batch.company_name = company_name
            batch.files.append(detail_path.name)
            batch.rows.extend(rows)
        return batches, warnings

    def _read_detail_rows(
        self,
        path: Path,
        detail_type: str,
        should_cancel: Callable[[], bool] | None = None,
    ) -> tuple[list[DetailPaymentRow], str, str]:
        raw = _read_detail_raw(path, f"расшифровку {DETAIL_TYPE_LABELS[detail_type]}")
        company_key, company_name = _company_key_from_detail(raw)
        aliases = DETAIL_ALIASES[detail_type]
        header_row = _detect_detail_header_row(raw, aliases, DETAIL_TYPE_LABELS[detail_type], path.name)
        headers = _build_headers(raw.iloc[header_row])
        df = raw.iloc[header_row + 1 :].copy()
        df.columns = headers
        df = df.dropna(how="all")

        columns = DetailColumns(
            amount=_find_column(headers, "amount", aliases),
            iin=_find_column(headers, "iin", aliases),
            fio=_find_column(headers, "fio", aliases),
        )
        missing_required = []
        if not columns.amount:
            missing_required.append("Сумма")
        if not columns.iin:
            missing_required.append(_detail_iin_column_label(detail_type))
        if missing_required:
            raise ProcessingError(
                f"В расшифровке {path.name} не найдены обязательные колонки: "
                + ", ".join(missing_required)
            )

        rows: list[DetailPaymentRow] = []
        for row_index, (dataframe_index, row) in enumerate(df.iterrows()):
            if row_index % 100 == 0:
                _check_cancelled(should_cancel)
            amount = _parse_amount(row.get(columns.amount)) if columns.amount else None
            if not _is_positive_amount(amount):
                continue

            iin = _normalize_iin(row.get(columns.iin)) if columns.iin else ""
            if not iin:
                continue

            rows.append(
                DetailPaymentRow(
                    detail_type=detail_type,
                    source_file=path.name,
                    source_row=int(dataframe_index) + 1,
                    iin=iin,
                    fio=_clean_cell(row.get(columns.fio)) if columns.fio else "",
                    amount=amount,
                )
            )

        return rows, company_key, company_name

    def _build_records(
        self,
        table: NewStatementTable,
        should_cancel: Callable[[], bool] | None = None,
    ) -> tuple[list[NewPaymentRecord], int, int, list[dict[str, Any]]]:
        records: list[NewPaymentRecord] = []
        skipped_rows: list[dict[str, Any]] = []
        skipped_amount_rows = 0
        skipped_excluded_rows = 0
        columns = table.columns

        for row_index, (dataframe_index, row) in enumerate(table.df.iterrows()):
            if row_index % 100 == 0:
                _check_cancelled(should_cancel)
            if _is_statement_summary_row(row, columns.amount):
                skipped_excluded_rows += 1
                skipped_rows.append(
                    _skipped_statement_row(
                        table,
                        row,
                        dataframe_index,
                        "Итоговая строка выписки, не является отдельным платежом",
                    )
                )
                continue

            if _is_amount_only_row(row, columns.amount):
                skipped_excluded_rows += 1
                skipped_rows.append(
                    _skipped_statement_row(
                        table,
                        row,
                        dataframe_index,
                        "Строка содержит только итоговую сумму и не является отдельным платежом",
                    )
                )
                continue

            sender = _first_non_empty(
                _clean_cell(row.get(columns.sender)) if columns.sender else "",
                _clean_cell(row.get(columns.counterparty)) if columns.counterparty else "",
            )
            purpose = _first_non_empty(
                _clean_cell(row.get(columns.purpose)) if columns.purpose else "",
                _clean_cell(row.get(columns.counterparty)) if columns.counterparty else "",
                sender,
            )
            if _new_is_excluded_purpose(purpose, sender):
                skipped_excluded_rows += 1
                skipped_rows.append(
                    _skipped_statement_row(
                        table,
                        row,
                        dataframe_index,
                        "Платеж исключен правилом назначения/отправителя",
                    )
                )
                continue

            amount = _parse_amount(row.get(columns.amount)) if columns.amount else None
            if not _is_positive_amount(amount):
                skipped_amount_rows += 1
                skipped_rows.append(
                    _skipped_statement_row(
                        table,
                        row,
                        dataframe_index,
                        "Нет положительной суммы в колонке поступления (Кредит)",
                    )
                )
                continue

            account = _first_non_empty(
                _clean_cell(row.get(columns.account)) if columns.account else "",
                table.filename_account,
            )
            payment_method = _payment_method_from_sender(sender)
            counterparty_iin = _normalize_iin(row.get(columns.iin)) if columns.iin else ""

            records.append(
                NewPaymentRecord(
                    date=_parse_date(row.get(columns.date)) if columns.date else None,
                    sender=sender,
                    amount=amount,
                    purpose=purpose,
                    payment_method=payment_method,
                    account=account,
                    company=table.filename_company,
                    counterparty_iin=counterparty_iin,
                    source_row=int(dataframe_index) + 1,
                    source_file=table.source_file,
                )
            )

        return records, skipped_amount_rows, skipped_excluded_rows, skipped_rows

    def _find_match(self, record: NewPaymentRecord, reference: NewReferenceBook) -> NewMatch:
        purpose_match = _find_in_new_text(
            record,
            reference,
            record.purpose,
            "Назначение платежа",
            record.purpose_names,
        )
        if purpose_match.entry or _is_blocking_new_match(purpose_match):
            return purpose_match
        if record.purpose_names:
            # An explicit person in the purpose has priority over the sender.
            # Falling back to a different sender name could assign a payment
            # made on behalf of a debtor to the payer instead.
            return purpose_match

        sender_text = "\n".join(part for part in (record.counterparty_iin, record.sender) if part)
        sender_names = (
            record.sender_names
            if record.payment_method == "Физическое лицо" and not _is_chsi_sender(record.sender)
            else ()
        )
        sender_match = _find_in_new_text(
            record,
            reference,
            sender_text,
            "Отправитель",
            sender_names,
        )
        if sender_match.entry or _is_blocking_new_match(sender_match):
            return sender_match

        combined_text = "\n".join(
            part for part in (record.counterparty_iin, record.purpose, record.sender) if part
        )
        return _find_in_new_text(
            record,
            reference,
            combined_text,
            "Назначение платежа + Отправитель",
            tuple(dict.fromkeys((*record.purpose_names, *sender_names))),
        )

    def _append_statement_record(
        self,
        record: NewPaymentRecord,
        reference: NewReferenceBook,
        matched_rows: list[dict[str, Any]],
        not_found_rows: list[dict[str, Any]],
    ) -> None:
        match = self._find_match(record, reference)
        base_row = _base_output_row(record, record.amount)

        special_reason = _new_special_not_found_reason(record)
        if special_reason:
            not_found_rows.append(
                {
                    "ДБЗ": "",
                    "ИИН": record.counterparty_iin,
                    "ФИО (заёмщика)": "",
                    **base_row,
                    "Взыскатель": "",
                    "ФИО из выписки": "",
                    "Исходная строка": record.source_row,
                    "Причина": special_reason,
                }
            )
            return

        if match.entry:
            matched_rows.append(_matched_output_row(match.entry, base_row))
            return

        not_found_rows.append(
            {
                "ДБЗ": "",
                "ИИН": match.detected_iin,
                "ФИО (заёмщика)": "",
                **base_row,
                "Взыскатель": "",
                "ФИО из выписки": match.detected_name or "; ".join(record.purpose_names),
                "Исходная строка": record.source_row,
                "Причина": _format_new_reason(match),
            }
        )

    def _append_detail_batch(
        self,
        detail_type: str,
        batch: DetailBatch,
        statement_records: list[NewPaymentRecord],
        reference: NewReferenceBook,
        matched_rows: list[dict[str, Any]],
        not_found_rows: list[dict[str, Any]],
        reconciliation_errors: list[dict[str, Any]],
        should_cancel: Callable[[], bool] | None = None,
    ) -> int:
        statement_total = _round_money(sum(record.amount for record in statement_records))
        detail_total = batch.total_amount
        context = statement_records[0]

        if detail_total > statement_total + MONEY_TOLERANCE:
            reconciliation_errors.append(
                _reconciliation_error_row(
                    detail_type=detail_type,
                    batch=batch,
                    statement_records=statement_records,
                    reason="Сумма расшифровки больше суммы выписки. Записи расшифровки проигнорированы.",
                )
            )
            if detail_type in DETAIL_TYPES_IGNORE_OVERPAYMENT:
                return 0
            not_found_rows.append(
                {
                    "ДБЗ": "",
                    "ИИН": "",
                    "ФИО (заёмщика)": "",
                    **_base_output_row(context, statement_total),
                    "Взыскатель": "",
                    "ФИО из выписки": "",
                    "Исходная строка": ", ".join(batch.files),
                    "Причина": (
                        "Сумма расшифровки больше суммы выписки. "
                        "Платеж выписки отправлен в Не найдено, строки расшифровки не загружены в реестр."
                    ),
                }
            )
            return 0

        added_rows = 0
        for row_index, detail_row in enumerate(batch.rows):
            if row_index % 100 == 0:
                _check_cancelled(should_cancel)
            entry, reason = _match_detail_row_by_iin(detail_row.iin, reference)
            base_row = _base_output_row(context, detail_row.amount)
            if reason == "not_found":
                not_found_rows.append(
                    {
                        "ДБЗ": "",
                        "ИИН": detail_row.iin,
                        "ФИО (заёмщика)": "",
                        **base_row,
                        "Взыскатель": "",
                        "ФИО из выписки": detail_row.fio,
                        "Исходная строка": f"{detail_row.source_file}:{detail_row.source_row}",
                        "Причина": "ИИН из расшифровки не найден в справочнике",
                    }
                )
                continue

            if reason == "ambiguous":
                not_found_rows.append(
                    {
                        "ДБЗ": "",
                        "ИИН": detail_row.iin,
                        "ФИО (заёмщика)": "",
                        **base_row,
                        "Взыскатель": "",
                        "ФИО из выписки": detail_row.fio,
                        "Исходная строка": f"{detail_row.source_file}:{detail_row.source_row}",
                        "Причина": "По ИИН найдено два и более ДБЗ",
                    }
                )
                continue

            if entry:
                matched_rows.append(_matched_output_row(entry, base_row))
                added_rows += 1

        if detail_total + MONEY_TOLERANCE < statement_total:
            not_found_rows.append(
                {
                    "ДБЗ": "",
                    "ИИН": "",
                    "ФИО (заёмщика)": "",
                    **_base_output_row(context, _round_money(statement_total - detail_total)),
                    "Взыскатель": "",
                    "ФИО из выписки": "",
                    "Исходная строка": ", ".join(batch.files),
                    "Причина": (
                        "Сумма расшифровки меньше суммы выписки. "
                        "В строку вынесена разница между выпиской и расшифровкой."
                    ),
                }
            )

        return added_rows

    def _write_registry(
        self,
        path: Path,
        matched_rows: list[dict[str, Any]],
        not_found_rows: list[dict[str, Any]],
        reconciliation_errors: list[dict[str, Any]],
        skipped_rows: list[dict[str, Any]],
    ) -> None:
        registry_df = pd.DataFrame(matched_rows, columns=NEW_OUTPUT_COLUMNS)
        not_found_df = pd.DataFrame(not_found_rows, columns=NEW_NOT_FOUND_COLUMNS)
        reconciliation_df = pd.DataFrame(
            reconciliation_errors,
            columns=NEW_RECONCILIATION_COLUMNS,
        )
        skipped_df = pd.DataFrame(skipped_rows, columns=SKIPPED_ROW_COLUMNS)
        summary_df = _build_company_summary_df(
            matched_rows,
            not_found_rows,
            registry_company_column="\u043a\u043e\u043c\u043f\u0430\u043d\u0438\u044f",
            registry_amount_column="\u0421\u0443\u043c\u043c\u0430",
            not_found_company_column="\u043a\u043e\u043c\u043f\u0430\u043d\u0438\u044f",
            not_found_amount_column="\u0421\u0443\u043c\u043c\u0430",
        )
        with pd.ExcelWriter(path, engine="openpyxl", date_format="DD.MM.YYYY") as writer:
            registry_df.to_excel(writer, sheet_name="Оплата", index=False)
            not_found_df.to_excel(writer, sheet_name="Не найдено", index=False)
            reconciliation_df.to_excel(writer, sheet_name="Ошибки сверки", index=False)
            skipped_df.to_excel(writer, sheet_name="Журнал пропусков", index=False)
            summary_df.to_excel(writer, sheet_name="\u0421\u0432\u043e\u0434 \u043f\u043e \u043a\u043e\u043c\u043f\u0430\u043d\u0438\u044f\u043c", index=False)
            _format_workbook(writer.book)

    def _write_not_found(self, path: Path, not_found_rows: list[dict[str, Any]]) -> None:
        not_found_df = pd.DataFrame(not_found_rows, columns=NEW_NOT_FOUND_COLUMNS)
        with pd.ExcelWriter(path, engine="openpyxl", date_format="DD.MM.YYYY") as writer:
            not_found_df.to_excel(writer, sheet_name="Не найдено", index=False)
            _format_workbook(writer.book)


def _detect_new_header_row(raw: pd.DataFrame) -> int:
    best_row: int | None = None
    best_score = -1
    max_rows = min(len(raw), 120)
    for row_index in range(max_rows):
        normalized_cells = [_normalize_name(value) for value in raw.iloc[row_index].tolist()]
        if not any(_is_credit_header_cell(cell) for cell in normalized_cells):
            continue
        score = 0
        for aliases in NEW_STATEMENT_ALIASES.values():
            if any(any(alias == cell or alias in cell for alias in aliases) for cell in normalized_cells):
                score += 1
        if score > best_score:
            best_row = row_index
            best_score = score
    if best_row is None:
        raise ProcessingError(
            "В выписке нового формата не найдена строка заголовков с колонкой «Кредит»."
        )
    return best_row


def _detect_reference_header_row(raw: pd.DataFrame) -> int:
    best_row = 0
    best_score = -1
    max_rows = min(len(raw), 120)
    for row_index in range(max_rows):
        normalized_cells = [_normalize_name(value) for value in raw.iloc[row_index].tolist()]
        score = 0
        for aliases in NEW_REFERENCE_ALIASES.values():
            if any(any(alias == cell or alias in cell for alias in aliases) for cell in normalized_cells):
                score += 1
        if score > best_score:
            best_row = row_index
            best_score = score
    return best_row


def _detect_detail_header_row(
    raw: pd.DataFrame,
    aliases: dict[str, list[str]],
    detail_label: str,
    filename: str,
) -> int:
    best_row: int | None = None
    best_score = -1
    max_rows = min(len(raw), 120)
    for row_index in range(max_rows):
        headers = _build_headers(raw.iloc[row_index])
        amount_col = _find_column(headers, "amount", aliases)
        iin_col = _find_column(headers, "iin", aliases)
        fio_col = _find_column(headers, "fio", aliases)
        score = int(bool(amount_col)) * 3 + int(bool(iin_col)) * 3 + int(bool(fio_col))
        if score > best_score:
            best_row = row_index
            best_score = score
        if amount_col and iin_col:
            return row_index

    if best_row is None:
        raise ProcessingError(f"В расшифровке {filename} не найдена строка заголовков.")

    raise ProcessingError(
        f"В расшифровке {filename} ({detail_label}) не найдена строка заголовков "
        "с колонками суммы и ИИН."
    )


def _read_detail_raw(path: Path, label: str) -> pd.DataFrame:
    raw = _read_excel_raw(path, label)
    if path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx"}:
        return raw

    formula_raw = _read_openpyxl_formula_raw(path)
    if formula_raw is None:
        return raw

    rows = max(len(raw), len(formula_raw))
    columns = max(len(raw.columns), len(formula_raw.columns))
    raw = raw.reindex(index=range(rows), columns=range(columns))
    formula_raw = formula_raw.reindex(index=range(rows), columns=range(columns))
    return raw.combine_first(formula_raw)


def _read_openpyxl_formula_raw(path: Path) -> pd.DataFrame | None:
    try:
        workbook = load_workbook(path, data_only=False, read_only=False)
    except Exception:
        return None

    try:
        for worksheet in workbook.worksheets:
            values: list[list[Any]] = []
            has_data = False
            for row in worksheet.iter_rows():
                row_values = [_formula_value_for_detail(cell.value) for cell in row]
                if any(not _is_detail_empty(value) for value in row_values):
                    has_data = True
                values.append(row_values)
            if has_data:
                return pd.DataFrame(values, dtype=object)
        return None
    finally:
        workbook.close()


def _formula_value_for_detail(value: Any) -> Any:
    if isinstance(value, str) and value.startswith("="):
        quoted_value = value.strip()
        match = re.fullmatch(r'=\s*"([^"]*)"', quoted_value)
        if match:
            return match.group(1)
    return value


def _is_detail_empty(value: Any) -> bool:
    if value is None:
        return True
    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        return False


def _detail_type_for_path(path: Path) -> str:
    compact_name = _compact_text(path.stem)
    if "astanaplat" in compact_name or "астанаплат" in compact_name:
        return "astana"
    if "евраз" in compact_name or "eurasian" in compact_name or "evraz" in compact_name:
        return "eurasian"
    if "казпочта" in compact_name or "kazpost" in compact_name or "kazpochta" in compact_name:
        return "kazpost"
    if (
        "народный" in compact_name
        or "narodny" in compact_name
        or "narodnyi" in compact_name
        or "halyk" in compact_name
        or "халык" in compact_name
    ):
        return "halyk"
    return ""


def _detail_type_for_record(record: NewPaymentRecord) -> str:
    sender = _compact_text(record.sender)
    purpose = _clean_cell(record.purpose).lower()
    if "astanaplat" in sender:
        return "astana"
    if "евразийскийбанк" in sender or "eurasianbank" in sender or "evrazbank" in sender:
        return "eurasian"
    if "казпочта" in sender:
        return "kazpost"
    if "9005981" in purpose:
        return "halyk"
    return ""


def _detail_batch_key_for_record(
    record: NewPaymentRecord,
    detail_type: str,
    detail_batches: dict[tuple[str, str], DetailBatch],
) -> tuple[str, str] | None:
    if not detail_type:
        return None

    statement_company_key, _ = _company_key_from_statement(record)
    exact_key = (detail_type, statement_company_key)
    if statement_company_key and exact_key in detail_batches:
        return exact_key

    fallback_key = (detail_type, "")
    if fallback_key in detail_batches:
        return fallback_key

    return None


def _detail_iin_column_label(detail_type: str) -> str:
    if detail_type == "astana":
        return "Договор/лицевой счет"
    if detail_type in {"kazpost", "eurasian"}:
        return "ИИН/БИН плательщика"
    return "ИИН"


def _company_key_from_statement(record: NewPaymentRecord) -> tuple[str, str]:
    text = "\n".join(part for part in (record.company, record.sender, record.purpose) if part)
    return _company_key_from_text(text)


def _company_key_from_detail(raw: pd.DataFrame) -> tuple[str, str]:
    values = []
    for value in raw.to_numpy().flatten().tolist():
        text = _clean_cell(value)
        if text:
            values.append(text)
    return _company_key_from_text("\n".join(values))


def _company_key_from_text(text: str) -> tuple[str, str]:
    normalized = _normalize_name(text)
    for company_key, aliases in COMPANY_ALIASES.items():
        for alias in aliases:
            normalized_alias = _normalize_name(alias)
            if normalized_alias and normalized_alias in normalized:
                return company_key, alias

    company_name = _extract_too_company_name(text)
    if company_name:
        return _normalize_company_key(company_name), company_name
    return "", ""


def _extract_too_company_name(text: str) -> str:
    patterns = (
        r'\bТОО\s+"([^"]+)"',
        r"\bТОО\s+([A-Za-zА-Яа-яЁё0-9 .«»\"-]{3,80})",
        r"\bТОВАРИЩЕСТВО\s+С\s+ОГРАНИЧЕННОЙ\s+ОТВЕТСТВЕННОСТЬЮ\s+\"([^\"]+)\"",
    )
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            value = _clean_cell(match.group(1))
            value = re.split(r"\bБИН\b|\bИИК\b|\bБИК\b|,|\n", value, maxsplit=1, flags=re.IGNORECASE)[0]
            return value.strip(" .«»\"-")
    return ""


def _normalize_company_key(value: str) -> str:
    normalized = _normalize_name(value)
    stopwords = {
        "тоо",
        "товарищество",
        "с",
        "ограниченной",
        "ответственностью",
        "коллекторское",
        "агентство",
    }
    tokens = [token for token in normalized.split() if token not in stopwords]
    return "".join(tokens)


def _match_detail_row_by_iin(
    iin: str,
    reference: NewReferenceBook,
) -> tuple[NewReferenceEntry | None, str]:
    entries = reference.by_iin.get(iin, [])
    if not entries:
        return None, "not_found"
    dbz_values = reference.dbz_by_iin.get(iin, set())
    if len(dbz_values) > 1:
        return None, "ambiguous"
    return entries[0], ""


def _skipped_statement_row(
    table: NewStatementTable,
    row: pd.Series,
    dataframe_index: Any,
    reason: str,
) -> dict[str, Any]:
    columns = table.columns
    sender = _first_non_empty(
        _clean_cell(row.get(columns.sender)) if columns.sender else "",
        _clean_cell(row.get(columns.counterparty)) if columns.counterparty else "",
    )
    purpose = _first_non_empty(
        _clean_cell(row.get(columns.purpose)) if columns.purpose else "",
        _clean_cell(row.get(columns.counterparty)) if columns.counterparty else "",
        sender,
    )
    return {
        "Файл выписки": table.source_file,
        "Исходная строка": int(dataframe_index) + 1,
        "Дата": _parse_date(row.get(columns.date)) if columns.date else None,
        "Отправитель": sender,
        "Сумма": _parse_amount(row.get(columns.amount)) if columns.amount else None,
        "Назначение платежа": purpose,
        "Причина пропуска": reason,
    }


def _base_output_row(record: NewPaymentRecord, amount: float) -> dict[str, Any]:
    return {
        "Дата": record.date,
        "Отправитель": record.sender,
        "Сумма": amount,
        "Назначение платежа": record.purpose,
        "Оплата через": record.payment_method,
        "Номерсчета": record.account,
        "компания": record.company,
    }


def _matched_output_row(
    entry: NewReferenceEntry,
    base_row: dict[str, Any],
) -> dict[str, Any]:
    return {
        "ДБЗ": entry.dbz,
        "ИИН": entry.iin,
        "ФИО (заёмщика)": _clean_output_fio(entry.fio),
        **base_row,
        "Взыскатель": entry.creditor,
    }


def _statement_reconciliation_error_reason(record: NewPaymentRecord) -> str:
    purpose = _clean_cell(record.purpose).lower()
    for keyword in STATEMENT_RECONCILIATION_ERROR_PURPOSE_KEYWORDS:
        if keyword in purpose:
            return f"Исключено из реестра по назначению платежа: {keyword}"
    return ""


def _statement_reconciliation_error_row(
    record: NewPaymentRecord,
    reason: str,
) -> dict[str, Any]:
    return {
        "Тип расшифровки": "Выписка",
        "Сумма выписки": record.amount,
        "Сумма расшифровки": "",
        "Разница": record.amount,
        "Файлы расшифровки": "",
        "Дата": record.date,
        "Отправитель": record.sender,
        "Назначение платежа": record.purpose,
        "Причина": reason,
    }


def _reconciliation_error_row(
    detail_type: str,
    batch: DetailBatch,
    statement_records: list[NewPaymentRecord],
    reason: str,
) -> dict[str, Any]:
    context = statement_records[0]
    statement_total = _round_money(sum(record.amount for record in statement_records))
    detail_total = batch.total_amount
    return {
        "Тип расшифровки": DETAIL_TYPE_LABELS[detail_type],
        "Сумма выписки": statement_total,
        "Сумма расшифровки": detail_total,
        "Разница": _round_money(detail_total - statement_total),
        "Файлы расшифровки": ", ".join(batch.files),
        "Дата": context.date,
        "Отправитель": context.sender,
        "Назначение платежа": context.purpose,
        "Причина": reason,
    }


def _compact_text(value: Any) -> str:
    return "".join(_normalize_name(value).split())


def _round_money(value: float) -> float:
    return round(float(value or 0), 2)


def _find_in_new_text(
    record: NewPaymentRecord,
    reference: NewReferenceBook,
    text: str,
    source: str,
    detected_names: tuple[str, ...] = (),
) -> NewMatch:
    search_text = "\n".join(part for part in (record.counterparty_iin, text) if part)
    candidates: dict[int, NewCandidate] = {}

    for iin in _extract_iins(search_text):
        for entry in reference.by_iin.get(iin, []):
            candidate = _new_candidate_for(candidates, entry)
            candidate.criteria.add("ИИН")
            candidate.detected_iin = iin

    for dbz_key in _extract_dbz_keys(search_text):
        for entry in reference.by_dbz.get(dbz_key, []):
            _new_candidate_for(candidates, entry).criteria.add("ДБЗ")

    for detected_name in detected_names:
        name_entries: dict[int, NewReferenceEntry] = {}
        for lookup_key in name_lookup_keys(detected_name):
            for entry in reference.by_name_token.get(lookup_key, []):
                name_entries[entry.row_id] = entry
        for entry in name_entries.values():
            evidence = match_person_name(entry.fio, detected_name)
            if not evidence:
                continue
            candidate = _new_candidate_for(candidates, entry)
            candidate.criteria.add(evidence.criterion)
            candidate.detected_name = detected_name
            candidate.name_strength = max(candidate.name_strength, evidence.strength)

    valid_candidates = list(candidates.values())
    if not valid_candidates:
        reason = (
            "ФИО, извлеченное GLiNER 2, не найдено в справочнике по безопасным правилам"
            if detected_names
            else "GLiNER 2 не обнаружил ФИО; ИИН и ДБЗ также не дали совпадения"
        )
        return NewMatch(
            entry=None,
            detected_name="; ".join(detected_names),
            reason=reason,
            source=source,
        )

    top_score = max(candidate.score for candidate in valid_candidates)
    top_candidates = [candidate for candidate in valid_candidates if candidate.score == top_score]
    if len(top_candidates) > 1 and all(
        not ({"ИИН", "ДБЗ"} & candidate.criteria)
        for candidate in top_candidates
    ):
        top_name_strength = max(candidate.name_strength for candidate in top_candidates)
        top_candidates = [
            candidate
            for candidate in top_candidates
            if candidate.name_strength == top_name_strength
        ]
    if len({_normalize_identifier(candidate.entry.dbz) for candidate in top_candidates}) > 1:
        best = top_candidates[0]
        shared_iin = any("ИИН" in candidate.criteria for candidate in top_candidates)
        return NewMatch(
            entry=None,
            detected_iin=best.detected_iin,
            detected_name=best.detected_name,
            criteria=tuple(sorted(best.criteria)),
            reason=(
                "По найденному ИИН/ФИО в справочнике найдено два или более ДБЗ"
                if shared_iin
                else "Совпадение ФИО GLiNER 2 неоднозначно: найдено два или более ДБЗ"
            ),
            source=source,
        )

    best = top_candidates[0]
    return NewMatch(
        entry=best.entry,
        detected_iin=best.detected_iin,
        detected_name=best.detected_name,
        criteria=tuple(sorted(best.criteria)),
        source=source,
    )


def _new_candidate_for(
    candidates: dict[int, NewCandidate],
    entry: NewReferenceEntry,
) -> NewCandidate:
    if entry.row_id not in candidates:
        candidates[entry.row_id] = NewCandidate(entry=entry)
    return candidates[entry.row_id]


def _best_new_candidate(candidates: Any) -> NewCandidate | None:
    candidate_list = list(candidates)
    if not candidate_list:
        return None
    return max(candidate_list, key=lambda candidate: candidate.score)


def _entry_tokens(entry: NewReferenceEntry) -> set[str]:
    tokens = name_lookup_keys(entry.fio)
    for keyword in entry.keywords:
        tokens.update(name_lookup_keys(keyword))
    return tokens


def _new_name_criterion(
    entry: NewReferenceEntry,
    normalized_text: str,
    text_tokens: set[str],
) -> str:
    if entry.normalized_fio and _phrase_in_text(entry.normalized_fio, normalized_text):
        return "ФИО"
    if entry.name_pair and _phrase_in_text(entry.name_pair, normalized_text):
        return "Фамилия имя"
    for keyword in entry.keywords:
        if keyword and _phrase_in_text(keyword, normalized_text):
            return "Ключевое слово"
    if entry.surname and entry.surname in text_tokens:
        return "Фамилия"
    return ""


def _is_blocking_new_match(match: NewMatch) -> bool:
    lowered_reason = match.reason.lower()
    return (
        "два или более дбз" in lowered_reason
        or "неоднознач" in lowered_reason
    )


def _format_new_reason(match: NewMatch) -> str:
    parts = [match.reason]
    if match.source:
        parts.append(f"поле: {match.source}")
    if match.criteria:
        parts.append("критерии: " + ", ".join(match.criteria))
    return "; ".join(parts)


def _new_special_not_found_reason(record: NewPaymentRecord) -> str:
    purpose = _clean_cell(record.purpose).lower()
    sender = _clean_cell(record.sender).lower()

    for keyword in NEW_PURPOSE_NOT_FOUND_KEYWORDS:
        if keyword in purpose:
            return f"Исключено из реестра по ключевому слову в назначении платежа: {keyword}"

    for keyword in NEW_SENDER_NOT_FOUND_KEYWORDS:
        if keyword in sender:
            return f"Исключено из реестра по отправителю: {keyword}"

    return ""


def _new_is_excluded_purpose(purpose: str, sender: str = "") -> bool:
    lowered = _clean_cell(purpose).lower()
    if "\u0432\u043e\u0437\u0432\u0440\u0430\u0442 \u043e\u0448\u0438\u0431\u043e\u0447\u043d\u044b\u0445 \u044d\u043b\u0435\u043c\u0435\u043d\u0442\u043e\u0432" in lowered:
        return True
    if (
        "\u043f\u0435\u0440\u0435\u0432\u043e\u0434 \u0441\u043e\u0431\u0441\u0442\u0432\u0435\u043d\u043d\u044b\u0445 \u0441\u0440\u0435\u0434\u0441\u0442\u0432" in lowered
        and _is_chsi_sender(sender)
    ):
        return False
    for keyword in NEW_EXCLUDED_PURPOSE_KEYWORDS:
        if not _contains_excluded_keyword(lowered, keyword):
            continue
        if keyword == "возврат" and _looks_like_repayment_return(purpose):
            continue
        return True
    return False


def _looks_like_repayment_return(purpose: str) -> bool:
    normalized = _normalize_name(purpose)
    if _extract_iins(purpose) or _extract_dbz_keys(purpose):
        return True
    repayment_markers = (
        "погашение",
        "заем",
        "займ",
        "кредит",
        "должник",
        "должнику",
        "уступки права требования",
        "цессии",
    )
    return any(marker in normalized for marker in repayment_markers)


def _is_chsi_sender(sender: str) -> bool:
    normalized = _normalize_name(sender)
    return (
        "\u0447\u0441\u0438" in normalized.split()
        or "\u0447\u0430\u0441\u0442\u043d\u044b\u0439 \u0441\u0443\u0434\u0435\u0431\u043d\u044b\u0439 \u0438\u0441\u043f\u043e\u043b\u043d\u0438\u0442\u0435\u043b\u044c" in normalized
    )


def _first_non_empty(*values: str) -> str:
    for value in values:
        if value:
            return value
    return ""


def _payment_method_from_sender(sender: str) -> str:
    text = _clean_cell(sender).lower()
    if not text:
        return ""

    normalized_text = _normalize_name(text)
    if _contains_payment_keyword(text, normalized_text, CHSI_PAYMENT_KEYWORDS):
        return "ЧСИ"

    if (
        _contains_payment_keyword(text, normalized_text, WITHHOLDING_PAYMENT_KEYWORDS)
        and not _is_astana_plat_sender(normalized_text)
    ):
        return "Удержание"

    return "Физическое лицо"


def _contains_payment_keyword(
    text: str,
    normalized_text: str,
    keywords: tuple[str, ...],
) -> bool:
    tokens = set(normalized_text.split())
    for keyword in keywords:
        normalized_keyword = _normalize_name(keyword)
        if not normalized_keyword:
            continue
        if " " in normalized_keyword:
            if normalized_keyword in normalized_text:
                return True
            continue
        if len(normalized_keyword) <= 4:
            if normalized_keyword in tokens:
                return True
            continue
        if normalized_keyword in normalized_text:
            return True
    return False


def _is_astana_plat_sender(normalized_text: str) -> bool:
    return "тоо" in normalized_text.split() and (
        "astana plat" in normalized_text
        or "астана плат" in normalized_text
    )


def _infer_payment_method(*values: str) -> str:
    text = " ".join(value.lower() for value in values if value)
    upper_text = " ".join(value.upper() for value in values if value)
    if "CTB" in upper_text:
        return "CTB"
    if "kaspi" in text or "каспи" in text:
        return "Kaspi"
    if "forte" in text or "форте" in text:
        return "ForteBank"
    if "halyk" in text or "народный" in text:
        return "Halyk"
    if "jusan" in text:
        return "Jusan"
    return ""
